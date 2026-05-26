"""One-off loader: ADGM CSV + Mauritius xlsx -> introducers table."""
import csv
import os
import sys
from openpyxl import load_workbook

from src.db import get_conn
from src.introducers import _normalise

ADGM_CSV = r"C:\Users\isuda\AppData\Local\Temp\ADGM CSPs List (1).csv"
MU_XLSX = r"C:\Users\isuda\AppData\Local\Temp\Mauritius Management Companies List (1).xlsx"


def load_adgm():
    rows = []
    with open(ADGM_CSV, "r", encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            name = (r.get("ADGM Licensed CSP name") or "").strip()
            if not name:
                continue
            rows.append({
                "company_name": name,
                "jurisdiction": "ADGM",
                "contact_name": (r.get("Contact name") or "").strip() or None,
                "contact_email": (r.get("Email") or "").strip() or None,
                "source": "manual_upload",
                "notes": "ADGM Licensed CSP",
            })
    return rows


def load_mauritius():
    wb = load_workbook(MU_XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    # detect header row
    header = None
    header_idx = 0
    for i, r in enumerate(rows[:10]):
        joined = " ".join(str(c or "") for c in r).lower()
        if "name" in joined and ("email" in joined or "contact" in joined or "company" in joined):
            header = [str(c or "").strip() for c in r]
            header_idx = i
            break
    if header is None:
        header = [str(c or "").strip() for c in rows[0]]
        header_idx = 0
    print("MU_HEADER", header, file=sys.stderr)

    def col(name_substr):
        for i, h in enumerate(header):
            if name_substr.lower() in h.lower():
                return i
        return None

    name_i = col("company") if col("company") is not None else col("name")
    if name_i is None:
        name_i = 0
    email_i = col("email")
    contact_i = col("contact person") or col("person")
    phone_i = col("contact number") or col("phone") or col("tel")
    addr_i = col("address")
    web_i = col("website")

    out = []
    for r in rows[header_idx + 1 :]:
        if not r or all(c is None or str(c).strip() == "" for c in r):
            continue
        nm = str(r[name_i] or "").strip()
        if not nm or nm.lower() in ("name", "company name", "#"):
            continue
        out.append({
            "company_name": nm,
            "jurisdiction": "Mauritius",
            "contact_email": (str(r[email_i]).strip() if email_i is not None and r[email_i] else None),
            "contact_name": (str(r[contact_i]).strip() if contact_i is not None and r[contact_i] else None),
            "phone_number": (str(r[phone_i]).strip() if phone_i is not None and r[phone_i] else None),
            "address": (str(r[addr_i]).strip() if addr_i is not None and r[addr_i] else None),
            "verify_url": (str(r[web_i]).strip() if web_i is not None and r[web_i] else None),
            "source": "manual_upload",
            "notes": "Mauritius Management Company",
        })
    return out


def upsert(rows):
    inserted = updated = 0
    with get_conn() as conn:
        with conn.cursor() as cur:
            for row in rows:
                cur.execute(
                    """
                    INSERT INTO introducers (
                        company_name, normalised_name, jurisdiction,
                        contact_email, contact_name, phone_number, address,
                        verify_url, source, notes
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT (normalised_name, jurisdiction) DO UPDATE SET
                        company_name = EXCLUDED.company_name,
                        contact_email = COALESCE(EXCLUDED.contact_email, introducers.contact_email),
                        contact_name = COALESCE(EXCLUDED.contact_name, introducers.contact_name),
                        phone_number = COALESCE(EXCLUDED.phone_number, introducers.phone_number),
                        address = COALESCE(EXCLUDED.address, introducers.address),
                        verify_url = COALESCE(EXCLUDED.verify_url, introducers.verify_url),
                        notes = EXCLUDED.notes,
                        updated_at = NOW()
                    RETURNING (xmax = 0) AS inserted
                    """,
                    (
                        row["company_name"],
                        _normalise(row["company_name"]),
                        row["jurisdiction"],
                        row.get("contact_email"),
                        row.get("contact_name"),
                        row.get("phone_number"),
                        row.get("address"),
                        row.get("verify_url"),
                        row.get("source") or "manual_upload",
                        row.get("notes"),
                    ),
                )
                is_new = cur.fetchone()[0]
                if is_new:
                    inserted += 1
                else:
                    updated += 1
            conn.commit()
    return inserted, updated


if __name__ == "__main__":
    adgm = load_adgm()
    mu = load_mauritius()
    print(f"ADGM rows: {len(adgm)}")
    print(f"MU rows: {len(mu)}")
    if mu[:3]:
        print("MU sample:", mu[:3])
    if "--apply" in sys.argv:
        ins, upd = upsert(adgm + mu)
        print(f"INSERTED {ins}  UPDATED {upd}")
    else:
        print("(dry run; pass --apply to write)")
