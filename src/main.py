import csv
import hmac
import io
import logging
import re
from datetime import date, datetime, timedelta, timezone
from html import escape
from urllib.parse import urlencode
from uuid import UUID

from fastapi import FastAPI, File, Form, HTTPException, Request, Response, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from itsdangerous import BadSignature, SignatureExpired, TimestampSigner
from psycopg.types.json import Jsonb
from pythonjsonlogger import jsonlogger

from src.config import (
    ACTOR_NAMES,
    ACTIVE_TERMINAL_STATUSES,
    ADMIN_ACTOR_NAMES,
    ADMIN_TOKEN,
    APP_ENV,
    CH_ENRICHMENT_SAFE_LIMIT,
    LOG_LEVEL,
    RM_NAMES,
    SCORING_DISPLAY_ENABLED,
    SCORING_SHADOW_MODE,
    SECRET_KEY,
    PROSPECT_ENGINE_DEMO_MODE,
    PROSPECT_ENGINE_FOOTER_TEXT,
    SHADOW_BACKFILL_BATCH_SIZE,
    SHADOW_BACKFILL_LOCK_TIMEOUT_MS,
    SHADOW_BACKFILL_MAX_BATCHES,
    SHADOW_SCORE_ACTIVE_STALE_DAYS,
)
from src.db import check_connection, get_conn
from src.domain.statuses import (
    normalize_status,
    require_canonical_status,
    status_label,
    status_options,
)
from src.ingestion.companies_house import run_ch_enrichment_batch
from src.ingestion.lei_backfill import backfill_lei_company_links
from src.route_intelligence import (
    CONTACTABILITY_LABELS,
    CONTACTABILITY_STATUS_LABELS,
    best_route_label,
    contactability_decision,
    contactability_status,
    contactability_status_label,
    lead_readiness,
    rm_status,
    rm_status_label,
    source_reliability_label,
    suggested_opener,
)
from src.scoring import SCORING_VERSION, SIGNAL_DETAILS
from src.security.write_auth import write_guard_required
from src.security.url_safety import sanitize_external_url
from src.shadow_scoring import backfill_active_shadow_scores

# --- Logging setup ---
handler = logging.StreamHandler()
formatter = jsonlogger.JsonFormatter("%(asctime)s %(name)s %(levelname)s %(message)s")
handler.setFormatter(formatter)
root_logger = logging.getLogger()
root_logger.handlers = [handler]
root_logger.setLevel(getattr(logging, LOG_LEVEL.upper(), logging.INFO))

logger = logging.getLogger(__name__)

_ACTOR_COOKIE = "actor"
_ACTOR_MAX_AGE = 30 * 24 * 3600
_actor_signer = TimestampSigner(SECRET_KEY)


def _read_actor(request: Request) -> str:
    raw = request.cookies.get(_ACTOR_COOKIE)
    if not raw:
        return ""
    try:
        unsigned = _actor_signer.unsign(raw, max_age=_ACTOR_MAX_AGE)
    except SignatureExpired:
        logger.info("actor_cookie_expired")
        return ""
    except BadSignature:
        logger.warning("actor_cookie_bad_signature")
        return ""
    return unsigned.decode("utf-8", errors="replace")


def _require_admin_token(request: Request) -> None:
    """Enforce bearer-token auth on admin routes. Constant-time compare;
    never logs the token value."""
    if not ADMIN_TOKEN:
        logger.error("admin_token_not_configured")
        raise HTTPException(status_code=503, detail="Admin not configured")
    header = request.headers.get("authorization", "")
    scheme, _, presented = header.partition(" ")
    if scheme.lower() != "bearer" or not presented:
        raise HTTPException(status_code=401, detail="Authentication required")
    if not hmac.compare_digest(presented, ADMIN_TOKEN):
        logger.warning("admin_token_mismatch")
        raise HTTPException(status_code=401, detail="Authentication required")


def _time_ago(dt: datetime | None) -> str:
    if dt is None:
        return "Never"
    now = datetime.now(timezone.utc)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    seconds = int((now - dt).total_seconds())
    if seconds < 120:
        return "Just now"
    if seconds < 3600:
        return f"{seconds // 60}m ago"
    if seconds < 86400:
        return f"{seconds // 3600}h ago"
    return f"{seconds // 86400}d ago"


app = FastAPI(
    title="Arie Leads",
    docs_url=None if APP_ENV == "production" else "/docs",
    redoc_url=None,
)
app.mount("/static", StaticFiles(directory="src/static"), name="static")
templates = Jinja2Templates(directory="src/templates")

from src.introducers import router as introducers_router  # noqa: E402
app.include_router(introducers_router)

_STATUS_OPTIONS = status_options()


def _canonical_status_or_422(raw_status: str) -> str:
    try:
        return require_canonical_status(raw_status)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc


def _format_entity_type(raw: str | None) -> str:
    if not raw:
        return "—"
    mapping = {
        "ltd": "Private Ltd",
        "limited": "Private Ltd",
        "private-limited-company": "Private Ltd",
        "private limited company": "Private Ltd",
        "plc": "PLC",
        "public-limited-company": "PLC",
        "llp": "LLP",
        "limited-liability-partnership": "LLP",
        "private-unlimited-company": "Unlimited",
        "private-limited-guarant-nsc": "Guarantee Co.",
        "registered-overseas-entity": "Overseas Entity",
        "uk-establishment": "UK Establishment",
    }
    return mapping.get(raw.lower().strip(), raw.title())


_UPLOAD_REQUIRED_COLUMNS = ["company_name", "jurisdiction"]
_UPLOAD_MAX_ROWS = 10000
_UPLOAD_MAX_BYTES = 10 * 1024 * 1024


def _build_query_string(params: dict) -> str:
    cleaned = {key: value for key, value in params.items() if value not in (None, "")}
    return urlencode(cleaned)


def _render_action_panel(lead_id: UUID, assigned_to: str, status: str, notes: str, contacted_at, follow_up_at, saved: bool = False) -> HTMLResponse:
    return HTMLResponse(
        f"""
        <div class="card" id="action-panel">
          <h2>RM Actions</h2>
          <form
            hx-post="/leads/{lead_id}/action"
            hx-target="#action-panel"
            hx-swap="outerHTML"
            style="display:flex; flex-direction:column; gap:.75rem"
          >
            <div>
              <label>Assign To</label>
              <select name="assigned_to">
                <option value="">— Unassigned —</option>
                {''.join(f'<option value="{escape(rm)}" {"selected" if rm == assigned_to else ""}>{escape(rm)}</option>' for rm in RM_NAMES)}
              </select>
            </div>
            <div>
              <label>Status</label>
              <select name="status">
                {''.join(
                    f'<option value="{escape(item["value"])}" {"selected" if item["value"] == status else ""}>{escape(item["label"])}</option>'
                    for item in _STATUS_OPTIONS
                )}
              </select>
            </div>
            <div>
              <label>Notes</label>
              <textarea name="notes">{escape(notes or "")}</textarea>
            </div>
            <div class="grid-2">
              <div>
                <label>Contacted At</label>
                <input type="date" name="contacted_at" value="{contacted_at.date().isoformat() if contacted_at else ""}">
              </div>
              <div>
                <label>Follow Up</label>
                <input type="date" name="follow_up_at" value="{follow_up_at.date().isoformat() if follow_up_at else ""}">
              </div>
            </div>
            <button type="submit" class="btn btn-primary">Save</button>
            {"<div style=\"color:#065f46; font-size:12px; margin-top:.25rem\">✓ Saved</div>" if saved else ""}
          </form>
        </div>
        """
    )


def _parse_upload_csv(file_bytes: bytes) -> tuple[list[dict], list[str], list[str]]:
    errors: list[str] = []
    columns: list[str] = []
    rows: list[dict] = []

    if len(file_bytes) > _UPLOAD_MAX_BYTES:
        return [], [], [f"File exceeds {_UPLOAD_MAX_BYTES // (1024 * 1024)} MB limit."]

    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        return [], [], ["CSV must be UTF-8 encoded."]

    reader = csv.DictReader(io.StringIO(text))
    columns = list(reader.fieldnames or [])
    missing = [column for column in _UPLOAD_REQUIRED_COLUMNS if column not in columns]
    if missing:
        errors.append("Missing required columns: " + ", ".join(missing))

    if not columns:
        return [], [], errors + ["CSV header row is missing."]

    for index, row in enumerate(reader, start=1):
        if index > _UPLOAD_MAX_ROWS:
            errors.append(f"CSV exceeds {_UPLOAD_MAX_ROWS} row limit.")
            break

        cleaned = {key: (value or "").strip() for key, value in row.items() if key}
        if not cleaned.get("company_name"):
            errors.append(f"Row {index}: company_name is required.")
        if not cleaned.get("jurisdiction"):
            errors.append(f"Row {index}: jurisdiction is required.")
        rows.append(cleaned)

    return columns, rows, errors


@app.get("/health")
def health(response: Response):
    db_ok = check_connection()

    queue_rows = 0
    queue_refreshed_at = None
    queue_fresh = False
    mauritius_last_seen = None
    last_pipeline_run = None
    shadow_counts = {"scored": 0, "unscored": 0, "failed": 0}

    if db_ok:
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT COUNT(*), MAX(refreshed_at) FROM queue_snapshot")
                    row = cur.fetchone()
                    if row:
                        queue_rows = row[0] or 0
                        queue_refreshed_at = row[1]
                        if queue_refreshed_at:
                            if queue_refreshed_at.tzinfo is None:
                                queue_refreshed_at = queue_refreshed_at.replace(tzinfo=timezone.utc)
                            age = datetime.now(timezone.utc) - queue_refreshed_at
                            queue_fresh = age.total_seconds() < 25 * 3600
                    cur.execute(
                        "SELECT MAX(updated_at) FROM companies WHERE source_system = 'mauritius_mns'"
                    )
                    mu_row = cur.fetchone()
                    if mu_row and mu_row[0] is not None:
                        ts = mu_row[0]
                        if ts.tzinfo is None:
                            ts = ts.replace(tzinfo=timezone.utc)
                        mauritius_last_seen = ts
                    cur.execute(
                        """
                        SELECT started_at, completed_at, status, uk_count, mu_count,
                               scores_count, queue_rows, duration_seconds, error
                        FROM pipeline_runs
                        ORDER BY started_at DESC
                        LIMIT 1
                        """
                    )
                    pr_row = cur.fetchone()
                    if pr_row:
                        started_at = pr_row[0]
                        completed_at = pr_row[1]
                        if started_at and started_at.tzinfo is None:
                            started_at = started_at.replace(tzinfo=timezone.utc)
                        if completed_at and completed_at.tzinfo is None:
                            completed_at = completed_at.replace(tzinfo=timezone.utc)
                        last_pipeline_run = {
                            "started_at": started_at.isoformat() if started_at else None,
                            "completed_at": completed_at.isoformat() if completed_at else None,
                            "status": pr_row[2],
                            "uk_count": pr_row[3],
                            "mu_count": pr_row[4],
                            "scores_count": pr_row[5],
                            "queue_rows": pr_row[6],
                            "duration_seconds": float(pr_row[7]) if pr_row[7] is not None else None,
                            "error": pr_row[8],
                        }
        except Exception as exc:
            logger.warning("health_queue_check_failed", extra={"error": str(exc)})

    if db_ok:
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        SELECT
                            COUNT(*) FILTER (WHERE score_state = 'scored') AS scored,
                            COUNT(*) FILTER (WHERE score_state = 'unscored') AS unscored,
                            COUNT(*) FILTER (WHERE score_state = 'failed') AS failed
                        FROM companies
                        """
                    )
                    shadow_row = cur.fetchone()
                    if shadow_row:
                        shadow_counts = {
                            "scored": shadow_row[0] or 0,
                            "unscored": shadow_row[1] or 0,
                            "failed": shadow_row[2] or 0,
                        }
        except Exception as exc:
            logger.warning("health_shadow_score_check_failed", extra={"error": str(exc)})

    if not db_ok:
        response.status_code = 503
    elif not queue_fresh and queue_rows > 0:
        response.status_code = 503

    return {
        "status": "ok" if (db_ok and queue_fresh) else "degraded",
        "db": "connected" if db_ok else "unreachable",
        "queue_rows": queue_rows,
        "queue_refreshed_at": queue_refreshed_at.isoformat() if queue_refreshed_at else None,
        "queue_fresh": queue_fresh,
        "mauritius_last_seen": mauritius_last_seen.isoformat() if mauritius_last_seen else None,
        "last_pipeline_run": last_pipeline_run,
        "scoring_version": SCORING_VERSION,
        "shadow_scoring_enabled": SCORING_SHADOW_MODE,
        "shadow_scoring_display_enabled": SCORING_DISPLAY_ENABLED,
        "shadow_score_counts": shadow_counts,
    }


@app.post("/admin/lei-backfill")
def admin_lei_backfill(request: Request):
    _require_admin_token(request)
    with get_conn() as conn:
        result = backfill_lei_company_links(conn)
    return result


@app.post("/admin/ch-enrichment")
def admin_ch_enrichment(request: Request, limit: int = None):
    _require_admin_token(request)
    actual_limit = limit if limit is not None else CH_ENRICHMENT_SAFE_LIMIT
    with get_conn() as conn:
        result = run_ch_enrichment_batch(conn, actual_limit)
    return result


@app.post("/admin/shadow-scoring/backfill")
def admin_shadow_scoring_backfill(request: Request):
    _require_admin_token(request)
    with get_conn() as conn:
        result = backfill_active_shadow_scores(
            conn,
            stale_days=SHADOW_SCORE_ACTIVE_STALE_DAYS,
            batch_size=SHADOW_BACKFILL_BATCH_SIZE,
            max_batches=SHADOW_BACKFILL_MAX_BATCHES,
            lock_timeout_ms=SHADOW_BACKFILL_LOCK_TIMEOUT_MS,
        )
    return {
        "shadow_scoring_enabled": SCORING_SHADOW_MODE,
        "display_enabled": SCORING_DISPLAY_ENABLED,
        **result,
    }


@app.post("/me")
def set_actor(request: Request, actor: str = Form("")):
    raw_referer = request.headers.get("referer", "")
    try:
        from urllib.parse import urlparse
        parsed = urlparse(raw_referer)
        same_origin = parsed.scheme in ("http", "https") and parsed.netloc == request.headers.get("host", "")
        redirect_to = raw_referer if same_origin else "/"
    except Exception:
        redirect_to = "/"
    response = RedirectResponse(url=redirect_to, status_code=303)
    if actor and actor in ACTOR_NAMES:
        signed = _actor_signer.sign(actor.encode("utf-8")).decode("ascii")
        response.set_cookie(
            _ACTOR_COOKIE,
            signed,
            max_age=_ACTOR_MAX_AGE,
            httponly=True,
            secure=(APP_ENV == "production"),
            samesite="lax",
        )
    elif not actor:
        response.delete_cookie(_ACTOR_COOKIE)
    # if actor provided but not in ACTOR_NAMES: silently ignore, don't set cookie
    return response


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request):
    actor = _read_actor(request)
    now = datetime.now(timezone.utc)
    stale_cutoff = now - timedelta(hours=36)

    with get_conn() as conn:
        with conn.cursor() as cur:
            # Panel 1 — pipeline freshness
            cur.execute("""
                SELECT started_at, completed_at, status,
                       uk_count, mu_count, scores_count, queue_rows, duration_seconds
                FROM pipeline_runs ORDER BY started_at DESC LIMIT 1
            """)
            last_run_row = cur.fetchone()

            cur.execute("""
                SELECT started_at, uk_count, mu_count, scores_count, queue_rows
                FROM pipeline_runs WHERE status = 'success'
                ORDER BY started_at DESC LIMIT 1
            """)
            last_success_row = cur.fetchone()

            cur.execute(
                "SELECT MAX(last_enriched_at) FROM companies"
                " WHERE source_system = 'companies_house'"
            )
            last_enriched_at = cur.fetchone()[0]

            cur.execute("SELECT MAX(last_seen) FROM lei_records")
            last_lei_seen = cur.fetchone()[0]

            # Panel 2 — lead volume + scoring
            cur.execute("""
                SELECT
                    COUNT(*) AS total,
                    COUNT(*) FILTER (WHERE created_at >= NOW() - INTERVAL '7 days')  AS last_7d,
                    COUNT(*) FILTER (WHERE created_at >= NOW() - INTERVAL '30 days') AS last_30d
                FROM companies
            """)
            vol_row = cur.fetchone()

            cur.execute("""
                SELECT
                    COUNT(*) FILTER (WHERE score BETWEEN 0  AND 39)  AS s0_39,
                    COUNT(*) FILTER (WHERE score BETWEEN 40 AND 59)  AS s40_59,
                    COUNT(*) FILTER (WHERE score BETWEEN 60 AND 79)  AS s60_79,
                    COUNT(*) FILTER (WHERE score BETWEEN 80 AND 100) AS s80_100,
                    COUNT(*) AS total_scored
                FROM lead_scores WHERE is_current = TRUE
            """)
            score_row = cur.fetchone()

            # Panel 3 — enrichment coverage
            cur.execute("""
                SELECT
                    (SELECT COUNT(*) FROM companies WHERE source_system = 'companies_house')                      AS total_uk,
                    (SELECT COUNT(*) FROM companies WHERE source_system = 'companies_house'
                                                    AND last_enriched_at IS NOT NULL)                           AS enriched_uk,
                    (SELECT COUNT(DISTINCT company_id) FROM company_officers)                                    AS with_officers,
                    (SELECT COUNT(DISTINCT company_id) FROM company_pscs)                                        AS with_pscs,
                    (SELECT COUNT(*) FROM lei_records)                                                           AS total_lei,
                    (SELECT COUNT(*) FROM lei_records WHERE company_id IS NOT NULL)                              AS linked_lei,
                    (SELECT COUNT(*) FROM companies WHERE source_system = 'mauritius_mns')                       AS total_mu
            """)
            cov_row = cur.fetchone()

            # Panel 4 — queue + workflow
            cur.execute(
                "SELECT status, COUNT(*) AS cnt FROM rm_actions GROUP BY status ORDER BY cnt DESC"
            )
            status_counts = cur.fetchall()

            cur.execute("""
                SELECT i.company_name, COUNT(ia.id) AS lead_count
                FROM introducers i
                JOIN introducer_actions ia ON ia.introducer_id = i.id
                GROUP BY i.id, i.company_name
                ORDER BY lead_count DESC LIMIT 5
            """)
            top_introducers = cur.fetchall()

            # Panel 5 — RM productivity metrics
            cur.execute("""
                SELECT
                    assigned_to,
                    COUNT(*)                                                            AS total_assigned,
                    COUNT(*) FILTER (WHERE contacted_at IS NOT NULL)                   AS contacted,
                    COUNT(*) FILTER (WHERE status IN ('Client','Qualified'))            AS converted,
                    COUNT(*) FILTER (WHERE follow_up_at IS NOT NULL
                                      AND follow_up_at >= CURRENT_DATE)                AS pending_followups,
                    COUNT(*) FILTER (WHERE follow_up_at IS NOT NULL
                                      AND follow_up_at < CURRENT_DATE
                                      AND status NOT IN ('Client','Closed - Not Fit',
                                                         'Not Fit','Archived'))        AS overdue_followups
                FROM rm_actions
                WHERE assigned_to IS NOT NULL
                GROUP BY assigned_to
                ORDER BY total_assigned DESC
            """)
            rm_productivity = cur.fetchall()

            cur.execute("""
                SELECT
                    COUNT(*)                                                            AS total_with_actions,
                    COUNT(*) FILTER (WHERE contacted_at IS NOT NULL)                   AS total_contacted,
                    COUNT(*) FILTER (WHERE status IN ('Client','Qualified'))            AS total_converted,
                    COUNT(*) FILTER (WHERE follow_up_at IS NOT NULL
                                      AND follow_up_at < CURRENT_DATE
                                      AND status NOT IN ('Client','Closed - Not Fit',
                                                         'Not Fit','Archived'))        AS total_overdue,
                    AVG(EXTRACT(EPOCH FROM (contacted_at - ra.created_at))/86400)
                        FILTER (WHERE contacted_at IS NOT NULL)                        AS avg_days_to_contact
                FROM rm_actions ra
            """)
            rm_summary = cur.fetchone()

            cur.execute("""
                WITH latest_routes AS (
                    SELECT DISTINCT ON (lead_id)
                           lead_id, contactability_bucket, status
                    FROM route_recommendations
                    WHERE status <> 'superseded'
                    ORDER BY lead_id, generated_at DESC
                )
                SELECT
                    COUNT(*) FILTER (WHERE contactability_bucket = 'ready_to_contact'),
                    COUNT(*) FILTER (WHERE contactability_bucket = 'route_via_introducer_csp'),
                    COUNT(*) FILTER (WHERE contactability_bucket = 'direct_candidate_found'),
                    COUNT(*) FILTER (WHERE contactability_bucket IN (
                        'management_company_route_likely', 'registry_evidence_only',
                        'needs_route_research'
                    )),
                    COUNT(*) FILTER (WHERE contactability_bucket = 'no_usable_route'),
                    COUNT(*) FILTER (WHERE status = 'accepted'),
                    COUNT(*) FILTER (WHERE status = 'rejected')
                FROM latest_routes
            """)
            route_metrics_row = cur.fetchone()

            # Top Opportunities — the few leads an RM should actually work now.
            # Deterministic ordering over actionable routes only; the readiness
            # gate is applied in Python to flag genuinely RM-ready rows.
            cur.execute("""
                SELECT c.company_name, c.jurisdiction, c.entity_type, c.id,
                       ls.priority_score, ls.tier,
                       rr.contactability_bucket, rr.best_route_type,
                       rr.best_route_value, rr.confidence, rr.next_action,
                       (rr.evidence_summary IS NOT NULL
                        AND jsonb_array_length(rr.evidence_summary) > 0) AS has_evidence
                FROM companies c
                JOIN lead_scores ls ON ls.company_id = c.id AND ls.is_current = TRUE
                JOIN LATERAL (
                    SELECT contactability_bucket, best_route_type, best_route_value,
                           confidence, next_action, evidence_summary
                    FROM route_recommendations
                    WHERE lead_id = c.id AND status <> 'superseded'
                    ORDER BY generated_at DESC
                    LIMIT 1
                ) rr ON TRUE
                WHERE rr.contactability_bucket IN (
                    'ready_to_contact', 'route_via_introducer_csp'
                )
                ORDER BY
                    CASE rr.contactability_bucket
                        WHEN 'ready_to_contact' THEN 0 ELSE 1 END,
                    CASE rr.confidence
                        WHEN 'high' THEN 0 WHEN 'medium' THEN 1 ELSE 2 END,
                    ls.priority_score DESC NULLS LAST
                LIMIT 10
            """)
            top_opportunity_rows = cur.fetchall()

    def _ts(ts: datetime | None) -> datetime | None:
        if ts is None:
            return None
        return ts if ts.tzinfo else ts.replace(tzinfo=timezone.utc)

    def _pct(num: int, denom: int) -> float:
        return round(num * 100 / denom, 1) if denom else 0.0

    top_opportunities = [
        {
            "id": r[3],
            "company_name": r[0],
            "jurisdiction": r[1],
            "entity_type": _format_entity_type(r[2]),
            "priority_score": r[4],
            "tier": r[5],
            "decision": contactability_decision(r[6]),
            "best_route": best_route_label(r[7], r[8]),
            "source_reliability_label": source_reliability_label(r[9]),
            "next_action": r[10],
            "is_rm_ready": _opp_gate["is_rm_ready"],
            "rm_status_label": rm_status_label(
                readiness=_opp_gate["readiness"], contactability_bucket=r[6]
            ),
        }
        for r in top_opportunity_rows
        for _opp_gate in [
            lead_readiness(
                recommendation={
                    "contactability_bucket": r[6],
                    "confidence": r[9],
                    "evidence_summary": [True] if r[11] else [],
                    "next_action": r[10],
                },
                tier=r[5],
            )
        ]
    ]

    last_run_at = _ts(last_run_row[0]) if last_run_row else None
    last_enriched = _ts(last_enriched_at)
    last_lei = _ts(last_lei_seen)

    sources = [
        {
            "name": "Nightly Pipeline",
            "detail": "UK · Mauritius · Scoring",
            "last_run": _time_ago(last_run_at),
            "last_run_at": last_run_at,
            "status": last_run_row[2] if last_run_row else "—",
            "counts": (
                f"UK {last_success_row[1] or 0} · "
                f"MU {last_success_row[2] or 0} · "
                f"Scores {last_success_row[3] or 0}"
            ) if last_success_row else "—",
            "stale": last_run_at is None or last_run_at < stale_cutoff,
        },
        {
            "name": "CH Enrichment",
            "detail": "Officers · PSCs",
            "last_run": _time_ago(last_enriched),
            "last_run_at": last_enriched,
            "status": "enriched" if last_enriched else "—",
            "counts": "",
            "stale": last_enriched is None or last_enriched < stale_cutoff,
        },
        {
            "name": "LEI Backfill",
            "detail": "GLEIF linkage",
            "last_run": _time_ago(last_lei),
            "last_run_at": last_lei,
            "status": "linked" if last_lei else "—",
            "counts": "",
            "stale": last_lei is None or last_lei < stale_cutoff,
        },
    ]

    total_uk   = cov_row[0] if cov_row else 0
    enriched_uk = cov_row[1] if cov_row else 0
    with_officers = cov_row[2] if cov_row else 0
    with_pscs  = cov_row[3] if cov_row else 0
    total_lei  = cov_row[4] if cov_row else 0
    linked_lei = cov_row[5] if cov_row else 0
    total_mu   = cov_row[6] if cov_row else 0

    return templates.TemplateResponse(
        request,
        "dashboard.html",
        {
            "actor": actor,
            # Panel 1
            "sources": sources,
            # Panel 2
            "total_leads":   vol_row[0] if vol_row else 0,
            "leads_7d":      vol_row[1] if vol_row else 0,
            "leads_30d":     vol_row[2] if vol_row else 0,
            "s0_39":         score_row[0] if score_row else 0,
            "s40_59":        score_row[1] if score_row else 0,
            "s60_79":        score_row[2] if score_row else 0,
            "s80_100":       score_row[3] if score_row else 0,
            "total_scored":  score_row[4] if score_row else 0,
            # Panel 3
            "total_uk":       total_uk,
            "pct_enriched_uk": _pct(enriched_uk, total_uk),
            "with_officers":  with_officers,
            "with_pscs":      with_pscs,
            "total_lei":      total_lei,
            "pct_lei_linked": _pct(linked_lei, total_lei),
            "total_mu":       total_mu,
            # Panel 4
            "status_counts":   status_counts,
            "top_introducers": top_introducers,
            # Panel 5 — RM productivity
            "rm_productivity": [
                {
                    "name": r[0],
                    "total_assigned": r[1],
                    "contacted": r[2],
                    "converted": r[3],
                    "pending_followups": r[4],
                    "overdue_followups": r[5],
                    "contact_rate": round(r[2] * 100 / r[1], 0) if r[1] else 0,
                    "conversion_rate": round(r[3] * 100 / r[1], 0) if r[1] else 0,
                }
                for r in rm_productivity
            ],
            "rm_summary": {
                "total_with_actions": rm_summary[0] if rm_summary else 0,
                "total_contacted": rm_summary[1] if rm_summary else 0,
                "total_converted": rm_summary[2] if rm_summary else 0,
                "total_overdue": rm_summary[3] if rm_summary else 0,
                "avg_days_to_contact": (
                    round(float(rm_summary[4]), 1)
                    if rm_summary and rm_summary[4] is not None else None
                ),
                "contact_rate": (
                    round(rm_summary[1] * 100 / rm_summary[0], 0)
                    if rm_summary and rm_summary[0] else 0
                ),
                "conversion_rate": (
                    round(rm_summary[2] * 100 / rm_summary[0], 0)
                    if rm_summary and rm_summary[0] else 0
                ),
            },
            "route_metrics": {
                "ready_to_contact": route_metrics_row[0] if route_metrics_row else 0,
                "via_introducer_csp": route_metrics_row[1] if route_metrics_row else 0,
                "direct_candidate": route_metrics_row[2] if route_metrics_row else 0,
                "needs_research": route_metrics_row[3] if route_metrics_row else 0,
                "no_usable_route": route_metrics_row[4] if route_metrics_row else 0,
                "accepted_routes": route_metrics_row[5] if route_metrics_row else 0,
                "rejected_routes": route_metrics_row[6] if route_metrics_row else 0,
            },
            # Simplified 4-value rollup (alias layer) over the rich route_metrics
            # above. direct_candidate folds into research_required, consistent
            # with contactability_status().
            "contactability_status_metrics": {
                "ready_to_contact": route_metrics_row[0] if route_metrics_row else 0,
                "route_via_introducer": route_metrics_row[1] if route_metrics_row else 0,
                "research_required": (
                    (route_metrics_row[2] + route_metrics_row[3])
                    if route_metrics_row else 0
                ),
                "no_compliant_route_found": (
                    route_metrics_row[4] if route_metrics_row else 0
                ),
            },
            "contactability_status_labels": CONTACTABILITY_STATUS_LABELS,
            "top_opportunities": top_opportunities,
            # Management-facing acquisition rollup. Built only from data we hold:
            # the contactability projection, the route review state, and RM
            # follow-up/contact counts. Meetings/replies/won are intentionally
            # absent — no field backs them yet (see RM workflow stage).
            "client_acquisition": {
                "rm_ready": route_metrics_row[0] if route_metrics_row else 0,
                "warm_introducer": route_metrics_row[1] if route_metrics_row else 0,
                "with_contact_route": (
                    (route_metrics_row[0] + route_metrics_row[1])
                    if route_metrics_row else 0
                ),
                "requires_research": (
                    (route_metrics_row[2] + route_metrics_row[3])
                    if route_metrics_row else 0
                ),
                "no_compliant_route": route_metrics_row[4] if route_metrics_row else 0,
                "accepted_routes": route_metrics_row[5] if route_metrics_row else 0,
                "overdue_followups": rm_summary[3] if rm_summary else 0,
                "contacted": rm_summary[1] if rm_summary else 0,
                "coverage_pct": _pct(
                    (
                        route_metrics_row[0] + route_metrics_row[1]
                        + route_metrics_row[2] + route_metrics_row[3]
                        + route_metrics_row[4]
                    )
                    if route_metrics_row else 0,
                    vol_row[0] if vol_row else 0,
                ),
            },
        }
    )


_TERMINAL_STATUS_KEYS = {
    key for raw in ACTIVE_TERMINAL_STATUSES if (key := normalize_status(raw))
}
_ACTIVE_STATUS_SQL = tuple(_TERMINAL_STATUS_KEYS or {"not_relevant", "not_fit"})
_ELIGIBLE_ROUTE_BUCKETS = (
    "ready_to_contact",
    "route_via_introducer_csp",
    "direct_candidate_found",
)
_TEAM_VISIBLE_STATUSES = ("sent_to_team", "in_progress", "follow_up")
_EMAIL_PUBLISHED_STATUSES = ("sent_to_team",)


def _is_admin_actor(actor: str) -> bool:
    return bool(actor and actor in ADMIN_ACTOR_NAMES)


def _route_kind(route_type: str | None, route_bucket: str | None) -> str:
    if route_bucket == "route_via_introducer_csp":
        return "Route via Introducer"
    if route_type in {"introducer", "csp", "management_company", "fiduciary"}:
        return "Route via Introducer"
    if route_type in {"direct", "fund_administrator"}:
        return "Direct"
    return "Needs Research"


def _lead_category(reason_codes: list[str], entity_type: str | None) -> str:
    codes = set(reason_codes or [])
    if "INVESTMENT_VEHICLE" in codes:
        return "Investment / fund"
    if "HOLDING_STRUCTURE" in codes or "FINANCIAL_SIC" in codes:
        return "Holding / finance"
    if "MAURITIUS_GBC" in codes:
        return "Mauritius GBC"
    if "MAURITIUS_AC" in codes:
        return "Authorised Company"
    return _format_entity_type(entity_type)


def _active_status_clause(alias: str = "ra") -> str:
    return f"COALESCE({alias}.status, 'new') <> ALL(%s)"


def _researched_route_clause() -> str:
    return """
        rr.contactability_bucket = ANY(%s)
        AND NULLIF(TRIM(rr.best_route_value), '') IS NOT NULL
        AND NULLIF(TRIM(rr.next_action), '') IS NOT NULL
        AND jsonb_array_length(COALESCE(rr.evidence_summary, '[]'::jsonb)) > 0
    """


def _route_optional_exprs(cur) -> dict[str, str]:
    cur.execute(
        """
        SELECT column_name
        FROM information_schema.columns
        WHERE table_name = 'route_recommendations'
          AND column_name = ANY(%s)
        """,
        (
            [
                "route_source_url",
                "route_source_label",
                "route_source_type",
                "route_last_checked_at",
            ],
        ),
    )
    present = {row[0] for row in cur.fetchall()}
    return {
        "route_source_url": (
            "route_source_url" if "route_source_url" in present else "NULL::text"
        ),
        "route_source_label": (
            "route_source_label" if "route_source_label" in present else "NULL::text"
        ),
        "route_source_type": (
            "route_source_type" if "route_source_type" in present else "NULL::text"
        ),
        "route_last_checked_at": (
            "route_last_checked_at"
            if "route_last_checked_at" in present
            else "NULL::timestamptz"
        ),
    }


def _optional_column_exprs(
    cur, table_name: str, column_defaults: dict[str, str]
) -> dict[str, str]:
    cur.execute(
        """
        SELECT column_name
        FROM information_schema.columns
        WHERE table_name = %s
          AND column_name = ANY(%s)
        """,
        (table_name, list(column_defaults)),
    )
    present = {row[0] for row in cur.fetchall()}
    return {
        column: column if column in present else default
        for column, default in column_defaults.items()
    }


def _prospect_rows(where_sql: str, params: list[object], limit: int = 120) -> list[dict]:
    with get_conn() as conn:
        with conn.cursor() as cur:
            opt = _route_optional_exprs(cur)
            sql = f"""
                SELECT
                    c.id, c.company_name, c.jurisdiction, c.entity_type,
                    qs.reason_summary, qs.reason_codes, qs.priority_score, qs.tier,
                    ra.assigned_to, COALESCE(ra.status, 'new'), ra.follow_up_at,
                    ra.notes,
                    rr.contactability_bucket, rr.best_route_type, rr.best_route_value,
                    rr.evidence_summary, rr.next_action, rr.confidence,
                    rr.route_source_url, rr.route_source_label, rr.route_source_type,
                    rr.route_last_checked_at, rr.generated_at,
                    c.verify_url, c.website, c.source_ref
                FROM queue_snapshot qs
                JOIN companies c ON c.id = qs.canonical_company_id
                LEFT JOIN rm_actions ra ON ra.company_id = c.id
                LEFT JOIN LATERAL (
                    SELECT contactability_bucket, best_route_type, best_route_value,
                           evidence_summary, next_action, confidence,
                           {opt["route_source_url"]} AS route_source_url,
                           {opt["route_source_label"]} AS route_source_label,
                           {opt["route_source_type"]} AS route_source_type,
                           {opt["route_last_checked_at"]} AS route_last_checked_at,
                           generated_at
                    FROM route_recommendations
                    WHERE lead_id = c.id AND status <> 'superseded'
                    ORDER BY generated_at DESC
                    LIMIT 1
                ) rr ON TRUE
                WHERE {where_sql}
                ORDER BY qs.priority_score DESC, c.incorporation_date DESC NULLS LAST,
                         c.company_name ASC
                LIMIT %s
            """
            cur.execute(sql, params + [limit])
            rows = cur.fetchall()

    rendered = []
    for row in rows:
        reason_codes = row[5] or []
        evidence = row[15] or []
        route_source = row[19] or row[18] or row[20] or "Source captured"
        has_contact_route = bool(row[14] and str(row[14]).strip())
        route_value = row[14] or "Research route before contact"
        status_key = normalize_status(row[9]) or "new"
        rendered.append(
            {
                "id": row[0],
                "name": row[1],
                "lead_type": _format_entity_type(row[3]),
                "jurisdiction": row[2],
                "category": _lead_category(reason_codes, row[3]),
                "why_relevant": row[4] or "Priority signals matched.",
                "contact_route": route_value,
                "has_contact_route": has_contact_route,
                "contact_route_kind": (
                    _route_kind(row[13], row[12])
                    if has_contact_route
                    else "Needs Research"
                ),
                "status": status_key,
                "status_label": status_label(status_key),
                "owner": row[8] or "",
                "follow_up_date": row[10],
                "notes": row[11] or "",
                "route_bucket": row[12] or "",
                "route_type": row[13] or "",
                "evidence": evidence,
                "evidence_text": "; ".join(str(item) for item in evidence[:3]),
                "next_action": row[16] or "",
                "confidence": row[17] or "",
                "source_url": sanitize_external_url(row[18]),
                "source_label": route_source,
                "source_type": row[20] or "",
                "last_checked_at": row[21],
                "route_generated_at": row[22],
                "verify_url": sanitize_external_url(row[23]),
                "website": sanitize_external_url(row[24]),
                "source_ref": row[25] or "",
                "score": row[6],
                "tier": row[7],
                "open_url": f"/leads/{row[0]}",
                "suggested_opening_angle": (
                    suggested_opener(
                        company_name=row[1],
                        entity_type=row[3],
                        jurisdiction=row[2],
                        contactability_bucket=row[12] or "needs_route_research",
                        best_route_value=route_value,
                    )
                    or "Draft a short RM note from the route evidence before contact."
                ),
            }
        )
    return rendered


def _enriched_route_rows(limit: int = 120) -> list[dict]:
    where_sql = f"""
        {_active_status_clause()}
        AND {_researched_route_clause()}
    """
    return _prospect_rows(where_sql, [list(_ACTIVE_STATUS_SQL), list(_ELIGIBLE_ROUTE_BUCKETS)], limit)


def _team_visible_prospect_rows(limit: int = 120) -> list[dict]:
    where_sql = f"""
        {_active_status_clause()}
        AND {_researched_route_clause()}
        AND (
            COALESCE(ra.status, 'new') = ANY(%s)
            OR NULLIF(TRIM(COALESCE(ra.assigned_to, '')), '') IS NOT NULL
        )
    """
    return _prospect_rows(
        where_sql,
        [
            list(_ACTIVE_STATUS_SQL),
            list(_ELIGIBLE_ROUTE_BUCKETS),
            list(_TEAM_VISIBLE_STATUSES),
        ],
        limit,
    )


def _published_email_rows(limit: int = 12) -> list[dict]:
    where_sql = f"""
        {_active_status_clause()}
        AND {_researched_route_clause()}
        AND COALESCE(ra.status, 'new') = ANY(%s)
    """
    return _prospect_rows(
        where_sql,
        [
            list(_ACTIVE_STATUS_SQL),
            list(_ELIGIBLE_ROUTE_BUCKETS),
            list(_EMAIL_PUBLISHED_STATUSES),
        ],
        limit,
    )


def _build_prospect_tabs(is_admin: bool) -> dict[str, dict]:
    enriched = _enriched_route_rows()
    team_visible = _team_visible_prospect_rows()
    intro = [
        row for row in enriched
        if row["route_bucket"] == "route_via_introducer_csp"
        or row["route_type"] in {"introducer", "csp", "management_company", "fiduciary"}
    ]
    team_intro = [row for row in team_visible if row["id"] in {item["id"] for item in intro}]
    assigned = [row for row in team_visible if row["owner"]]
    followups = [
        row
        for row in team_visible
        if row["follow_up_date"] or row["status"] == "follow_up"
    ]
    tabs = {
        "enriched": {"label": "Sent to Team", "rows": team_visible},
        "route-introducer": {"label": "Route via Introducer", "rows": team_intro},
        "introducers": {"label": "Introducers", "rows": []},
        "assigned": {"label": "Assigned to Team", "rows": assigned},
        "followups": {"label": "Follow-ups", "rows": followups},
    }
    if is_admin:
        this_week = _prospect_rows("1=1", [], 120)
        enriched_ids = {row["id"] for row in enriched}
        tabs = {
            "this-week": {"label": "This Week's Leads", "rows": this_week},
            "needs-enrichment": {
                "label": "Needs Enrichment",
                "rows": [
                    row
                    for row in this_week
                    if row["id"] not in enriched_ids and not row["has_contact_route"]
                ],
            },
            "enriched": {"label": "Enriched Leads", "rows": enriched},
            "introducers": {"label": "Introducers", "rows": []},
            "route-introducer": {"label": "Route via Introducer", "rows": intro},
            "assigned": {"label": "Assigned to Team", "rows": [row for row in enriched if row["owner"]]},
            "archived": {
                "label": "Rejected / Archived",
                "rows": _prospect_rows(_active_status_clause().replace("<> ALL", "= ANY"), [list(_ACTIVE_STATUS_SQL)], 80),
            },
        }
    return tabs


def _usable_introducers(limit: int = 80) -> list[dict]:
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT i.id, i.company_name, i.jurisdiction, i.entity_type,
                       i.contact_name, i.contact_email, i.phone_number,
                       i.verify_url, ia.assigned_to, COALESCE(ia.status, 'new')
                FROM introducers i
                LEFT JOIN introducer_actions ia ON ia.introducer_id = i.id
                WHERE (
                    NULLIF(TRIM(COALESCE(i.contact_email, '')), '') IS NOT NULL
                    OR NULLIF(TRIM(COALESCE(i.phone_number, '')), '') IS NOT NULL
                    OR NULLIF(TRIM(COALESCE(i.contact_name, '')), '') IS NOT NULL
                )
                ORDER BY i.company_name ASC
                LIMIT %s
                """,
                (limit,),
            )
            rows = cur.fetchall()
    return [
        {
            "id": row[0],
            "name": row[1],
            "lead_type": _format_entity_type(row[3]),
            "jurisdiction": row[2],
            "category": "Introducer",
            "why_relevant": "Usable introducer contact route captured.",
            "contact_route": row[5] or row[6] or row[4] or "Contact route captured",
            "contact_route_kind": "Route via Introducer",
            "status": normalize_status(row[9]) or "new",
            "status_label": status_label(row[9]),
            "owner": row[8] or "",
            "follow_up_date": None,
            "notes": "",
            "evidence_text": "Introducer record includes a usable contact route.",
            "next_action": "RM to verify fit and decide whether to use this route.",
            "source_label": "Introducer directory",
            "source_url": sanitize_external_url(row[7]),
            "open_url": f"/introducers/{row[0]}",
            "suggested_opening_angle": "Use only where the introducer relationship is appropriate and approved.",
        }
        for row in rows
    ]


def _weekly_email_leads(limit: int = 12) -> list[dict]:
    return _published_email_rows(limit)


def _prospect_counts(tabs: dict[str, dict], email_leads: list[dict]) -> dict[str, int]:
    return {
        "enriched": len(tabs.get("enriched", {}).get("rows", [])),
        "route_intro": len(tabs.get("route-introducer", {}).get("rows", [])),
        "introducers": len(tabs.get("introducers", {}).get("rows", [])),
        "assigned": len(tabs.get("assigned", {}).get("rows", [])),
        "followups": len(tabs.get("followups", {}).get("rows", [])),
        "email_eligible": len(email_leads),
    }


@app.get("/", response_class=HTMLResponse)
def prospect_engine(request: Request):
    actor = _read_actor(request)
    requested_view = request.query_params.get("view", "team")
    is_demo_admin = PROSPECT_ENGINE_DEMO_MODE and requested_view == "admin"
    is_admin = _is_admin_actor(actor) or is_demo_admin
    logger.info(
        "prospect_engine_view",
        extra={
            "actor": actor or "",
            "requested_view": requested_view,
            "admin_actor_names": ADMIN_ACTOR_NAMES,
            "is_admin": is_admin,
        },
    )
    if requested_view == "admin" and not is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")

    mode = "admin" if requested_view == "admin" and is_admin else "team"
    tabs = _build_prospect_tabs(is_admin=(mode == "admin"))
    tabs["introducers"]["rows"] = _usable_introducers()
    default_tab = "this-week" if mode == "admin" else "enriched"
    active_tab = request.query_params.get("tab", default_tab)
    if active_tab not in tabs:
        active_tab = default_tab
    email_leads = _weekly_email_leads()
    counts = _prospect_counts(tabs, email_leads)
    current_rows = tabs[active_tab]["rows"]

    return templates.TemplateResponse(
        request=request,
        name="prospect_engine.html",
        context={
            "mode": mode,
            "tabs": tabs,
            "active_tab": active_tab,
            "rows": current_rows,
            "counts": counts,
            "email_leads": email_leads,
            "actor_names": ACTOR_NAMES,
            "admin_actor_names": ADMIN_ACTOR_NAMES,
            "current_actor": actor,
            "is_admin": is_admin,
            "demo_mode": PROSPECT_ENGINE_DEMO_MODE,
            "rm_names": RM_NAMES,
            "statuses": _STATUS_OPTIONS,
            "footer_text": PROSPECT_ENGINE_FOOTER_TEXT,
            "week_label": date.today().strftime("Week of %d %b %Y"),
        },
    )


@app.get("/weekly-email", response_class=HTMLResponse)
def weekly_email_preview(request: Request):
    actor = _read_actor(request)
    email_leads = _weekly_email_leads()
    return templates.TemplateResponse(
        request=request,
        name="weekly_email_preview.html",
        context={
            "email_leads": email_leads,
            "counts": {"email_eligible": len(email_leads)},
            "footer_text": PROSPECT_ENGINE_FOOTER_TEXT,
            "week_label": date.today().strftime("Week of %d %b %Y"),
            "actor_names": ACTOR_NAMES,
            "current_actor": actor,
        },
    )


@app.get("/weekly-email/email-safe", response_class=HTMLResponse)
def weekly_email_safe_template(request: Request):
    email_leads = _weekly_email_leads()
    return templates.TemplateResponse(
        request=request,
        name="weekly_email_inline.html",
        context={
            "email_leads": email_leads,
            "footer_text": PROSPECT_ENGINE_FOOTER_TEXT,
            "week_label": date.today().strftime("Week of %d %b %Y"),
        },
    )


@app.get("/queue", response_class=HTMLResponse)
def queue(request: Request):
    filters = {
        "tier": request.query_params.get("tier", ""),
        "jurisdiction": request.query_params.get("jurisdiction", ""),
        "assigned_to": request.query_params.get("assigned_to", ""),
        "status": request.query_params.get("status", ""),
        "date_from": request.query_params.get("date_from", ""),
        "date_to": request.query_params.get("date_to", ""),
        "sort": request.query_params.get("sort", "score"),
        "route_bucket": request.query_params.get("route_bucket", ""),
        "named_route": request.query_params.get("named_route", ""),
        "introducer_match": request.query_params.get("introducer_match", ""),
    }
    try:
        page = max(int(request.query_params.get("page", 1)), 1)
    except ValueError:
        page = 1
    page_size = 50

    where_clauses = ["1=1"]
    params: list[object] = []
    if filters["tier"]:
        where_clauses.append("qs.tier = %s")
        params.append(filters["tier"])
    if filters["jurisdiction"]:
        where_clauses.append("c.jurisdiction = %s")
        params.append(filters["jurisdiction"])
    if filters["assigned_to"]:
        where_clauses.append("ra.assigned_to = %s")
        params.append(filters["assigned_to"])
    if filters["status"]:
        status_filter = normalize_status(filters["status"])
        if status_filter:
            where_clauses.append("ra.status = %s")
            params.append(status_filter)
    if filters["date_from"]:
        where_clauses.append("c.incorporation_date >= %s")
        params.append(filters["date_from"])
    if filters["date_to"]:
        where_clauses.append("c.incorporation_date <= %s")
        params.append(filters["date_to"])
    if filters.get("route_bucket") in CONTACTABILITY_LABELS:
        where_clauses.append(
            """(
                SELECT rr2.contactability_bucket FROM route_recommendations rr2
                WHERE rr2.lead_id = c.id AND rr2.status <> 'superseded'
                ORDER BY rr2.generated_at DESC LIMIT 1
            ) = %s"""
        )
        params.append(filters["route_bucket"])
    if filters.get("named_route") == "yes":
        where_clauses.append(
            """EXISTS (
                SELECT 1 FROM route_recommendations rr2
                WHERE rr2.lead_id = c.id
                  AND rr2.status <> 'superseded'
                  AND NULLIF(TRIM(rr2.best_route_value), '') IS NOT NULL
            )"""
        )
    if filters.get("introducer_match") == "yes":
        where_clauses.append(
            """EXISTS (
                SELECT 1 FROM introducer_matches im
                WHERE im.lead_id = c.id AND im.status IN ('pending', 'accepted')
            )"""
        )

    sort_sql = {
        "score": "qs.priority_score DESC, c.company_name ASC",
        "date": "c.incorporation_date DESC NULLS LAST, c.jurisdiction ASC, qs.priority_score DESC, c.company_name ASC",
        "date_asc": "c.incorporation_date ASC NULLS LAST, c.jurisdiction ASC, qs.priority_score DESC, c.company_name ASC",
        "name": "c.company_name ASC",
    }.get(filters["sort"], "c.incorporation_date DESC NULLS LAST, c.jurisdiction ASC, qs.priority_score DESC, c.company_name ASC")

    count_sql = f"""
        SELECT COUNT(*)
        FROM queue_snapshot qs
        JOIN companies c ON c.id = qs.canonical_company_id
        LEFT JOIN rm_actions ra ON ra.company_id = c.id
        WHERE {' AND '.join(where_clauses)}
    """
    query_sql = f"""
        SELECT
            c.id,
            c.company_name,
            c.jurisdiction,
            c.entity_type,
            c.incorporation_date,
            c.verify_url,
            qs.priority_score,
            qs.tier,
            qs.reason_summary,
            ra.assigned_to,
            ra.status,
            qs.refreshed_at,
            c.website,
            qs.reason_codes,
            ra.follow_up_at,
            rr.contactability_bucket,
            rr.best_route_type,
            rr.best_route_value,
            rr.confidence,
            rr.next_action,
            rr.route_status,
            rr.generated_at AS route_generated_at,
            rr.route_candidate_id
        FROM queue_snapshot qs
        JOIN companies c ON c.id = qs.canonical_company_id
        LEFT JOIN rm_actions ra ON ra.company_id = c.id
        LEFT JOIN LATERAL (
            SELECT contactability_bucket, best_route_type, best_route_value,
                   confidence, next_action, status AS route_status, generated_at,
                   route_candidate_id
            FROM route_recommendations
            WHERE lead_id = c.id AND status <> 'superseded'
            ORDER BY generated_at DESC
            LIMIT 1
        ) rr ON TRUE
        WHERE {' AND '.join(where_clauses)}
        ORDER BY {sort_sql}
        LIMIT %s OFFSET %s
    """

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(count_sql, params)
            total = cur.fetchone()[0]
            cur.execute(query_sql, params + [page_size, (page - 1) * page_size])
            rows = cur.fetchall()
            cur.execute("SELECT MAX(refreshed_at) FROM queue_snapshot")
            refreshed_at = cur.fetchone()[0]

    rendered_rows = [
        {
            "id": row[0],
            "company_name": row[1],
            "jurisdiction": row[2],
            "entity_type": _format_entity_type(row[3]),
            "incorporation_date": row[4],
            "verify_url": sanitize_external_url(row[5]),
            "priority_score": row[6],
            "tier": row[7],
            "reason_summary": row[8],
            "assigned_to": row[9],
            "status": normalize_status(row[10]) or "new",
            "status_label": status_label(row[10]),
            "refreshed_at": row[11],
            "website": sanitize_external_url(row[12]),
            "reason_codes": row[13] or [],
            "follow_up_at": row[14],
            "route_bucket": row[15],
            "contactability_label": CONTACTABILITY_LABELS.get(row[15] or "", ""),
            "contactability_status": contactability_status(row[15]) if row[15] else "",
            "contactability_status_label": (
                contactability_status_label(row[15]) if row[15] else ""
            ),
            "best_route_type": row[16],
            "best_route_value": row[17],
            "route_confidence": row[18],
            "route_next_action": row[19],
            "route_status": row[20],
            "route_generated_at": row[21],
            "route_candidate_id": row[22],
        }
        for row in rows
    ]
    total_pages = max((total + page_size - 1) // page_size, 1)
    query_params = {key: value for key, value in filters.items() if value}

    return templates.TemplateResponse(
        request=request,
        name="queue.html",
        context={
            "rows": rendered_rows,
            "total": total,
            "refreshed_at": f"{refreshed_at.day} {refreshed_at.strftime('%b %Y, %H:%M')} UTC" if refreshed_at else None,
            "filters": filters,
            "rm_names": RM_NAMES,
            "statuses": _STATUS_OPTIONS,
            "page": page,
            "total_pages": total_pages,
            "query_string": _build_query_string(query_params),
            "actor_names": ACTOR_NAMES,
            "current_actor": (_read_actor(request) or ""),
            "now_date": date.today(),
            "contactability_labels": CONTACTABILITY_LABELS,
        },
    )


@app.get("/leads/export")
def export_leads(request: Request):
    """Export the current filtered queue view as an Excel file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    filters = {
        "tier": request.query_params.get("tier", ""),
        "jurisdiction": request.query_params.get("jurisdiction", ""),
        "assigned_to": request.query_params.get("assigned_to", ""),
        "status": request.query_params.get("status", ""),
        "date_from": request.query_params.get("date_from", ""),
        "date_to": request.query_params.get("date_to", ""),
        "sort": request.query_params.get("sort", "score"),
    }

    where_clauses = ["1=1"]
    params: list[object] = []
    if filters["tier"]:
        where_clauses.append("qs.tier = %s")
        params.append(filters["tier"])
    if filters["jurisdiction"]:
        where_clauses.append("c.jurisdiction = %s")
        params.append(filters["jurisdiction"])
    if filters["assigned_to"]:
        where_clauses.append("ra.assigned_to = %s")
        params.append(filters["assigned_to"])
    if filters["status"]:
        status_filter = normalize_status(filters["status"])
        if status_filter:
            where_clauses.append("ra.status = %s")
            params.append(status_filter)
    if filters["date_from"]:
        where_clauses.append("c.incorporation_date >= %s")
        params.append(filters["date_from"])
    if filters["date_to"]:
        where_clauses.append("c.incorporation_date <= %s")
        params.append(filters["date_to"])

    sort_sql = {
        "score": "qs.priority_score DESC, c.company_name ASC",
        "date": "c.incorporation_date DESC NULLS LAST, c.jurisdiction ASC, qs.priority_score DESC, c.company_name ASC",
        "date_asc": "c.incorporation_date ASC NULLS LAST, c.jurisdiction ASC, qs.priority_score DESC, c.company_name ASC",
        "name": "c.company_name ASC",
    }.get(filters["sort"], "c.incorporation_date DESC NULLS LAST, c.jurisdiction ASC, qs.priority_score DESC, c.company_name ASC")

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT
                    c.company_name,
                    qs.priority_score,
                    qs.tier,
                    ra.status,
                    ra.assigned_to,
                    c.jurisdiction,
                    c.entity_type,
                    c.incorporation_date,
                    c.website,
                    qs.reason_summary,
                    lr.lei_code,
                    ra.follow_up_at,
                    c.verify_url
                FROM queue_snapshot qs
                JOIN companies c ON c.id = qs.canonical_company_id
                LEFT JOIN rm_actions ra ON ra.company_id = c.id
                LEFT JOIN lei_records lr ON lr.company_id = c.id
                WHERE {' AND '.join(where_clauses)}
                ORDER BY {sort_sql}
                """,
                params,
            )
            rows = cur.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lead Queue"

    header_fill = PatternFill("solid", fgColor="1A0B3D")
    header_font = Font(bold=True, color="C9A44A", size=11)
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=False)

    headers = [
        "Company Name", "Score", "Tier", "Status", "Assigned RM",
        "Jurisdiction", "Entity Type", "Incorporation Date", "Website",
        "Primary Reason", "LEI", "Follow-up Date", "Registry Link",
    ]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Column widths
    col_widths = [38, 8, 10, 18, 18, 14, 22, 18, 35, 55, 24, 16, 40]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    for row in rows:
        company_name, score, tier, status_raw, assigned_to, jurisdiction, \
            entity_type, inc_date, website, reason_summary, lei_code, \
            follow_up_at, verify_url = row
        ws.append([
            company_name or "",
            score or 0,
            tier or "",
            status_label(status_raw) if status_raw else "New",
            assigned_to or "",
            jurisdiction or "",
            _format_entity_type(entity_type) or "",
            inc_date.isoformat() if inc_date else "",
            sanitize_external_url(website) or "",
            reason_summary or "",
            lei_code or "",
            follow_up_at.isoformat() if follow_up_at else "",
            sanitize_external_url(verify_url) or "",
        ])

    # Freeze header row
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"lead-queue-{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/leads/{lead_id}", response_class=HTMLResponse)
def lead_detail(request: Request, lead_id: UUID):
    with get_conn() as conn:
        with conn.cursor() as cur:
            rm_opt = _optional_column_exprs(
                cur,
                "rm_actions",
                {
                    "next_action": "NULL::text",
                    "next_action_due_date": "NULL::date",
                    "feedback": "NULL::text",
                    "feedback_note": "NULL::text",
                },
            )
            cur.execute(
                f"""
                SELECT c.id, c.company_name, c.jurisdiction, c.entity_type,
                       c.incorporation_date, c.registered_address, c.source_system,
                       c.source_ref, c.verify_url, c.website,
                       ls.score, ls.tier, ls.reason_codes, ls.reason_summary, ls.scoring_version,
                       ra.assigned_to, ra.status, ra.notes, ra.contacted_at, ra.follow_up_at,
                       {rm_opt["next_action"]}, {rm_opt["next_action_due_date"]},
                       {rm_opt["feedback"]}, {rm_opt["feedback_note"]}
                FROM companies c
                LEFT JOIN lead_scores ls ON ls.company_id = c.id AND ls.is_current = TRUE
                LEFT JOIN rm_actions ra ON ra.company_id = c.id
                WHERE c.id = %s
                """,
                (lead_id,),
            )
            row = cur.fetchone()
            if row is None:
                raise HTTPException(status_code=404, detail="Lead not found")

            cur.execute(
                """
                SELECT entity_type, entity_id, action, actor, old_value, new_value, created_at
                FROM audit_log
                WHERE entity_type = 'company' AND entity_id = %s
                ORDER BY created_at DESC
                LIMIT 50
                """,
                (lead_id,),
            )
            audit_rows = cur.fetchall()

            cur.execute(
                """
                SELECT lei_code, entity_status, registration_status,
                       registered_on, managing_lou, gleif_url, last_seen
                FROM lei_records
                WHERE company_id = %s
                """,
                (lead_id,),
            )
            lei_row = cur.fetchone()

            cur.execute(
                """
                SELECT officer_name, role, appointed_on, resigned_on,
                       nationality, country_of_residence
                FROM company_officers
                WHERE company_id = %s
                ORDER BY appointed_on DESC NULLS LAST, officer_name
                """,
                (lead_id,),
            )
            officers_rows = cur.fetchall()

            cur.execute(
                """
                SELECT name, kind, nationality, country_of_residence,
                       natures_of_control, notified_on, ceased_on
                FROM company_pscs
                WHERE company_id = %s
                ORDER BY notified_on DESC NULLS LAST, name
                """,
                (lead_id,),
            )
            pscs_rows = cur.fetchall()

            contact_opt = _optional_column_exprs(
                cur,
                "lead_contacts",
                {
                    "email_confidence": "NULL::text",
                    "linkedin_verified": "FALSE",
                    "is_decision_maker": "FALSE",
                    "contact_priority": "(0)::int",
                    "enrichment_status": "'manual'::text",
                },
            )
            cur.execute(
                """
                SELECT id, name, role, email, phone, linkedin_url, source, notes,
                       created_at,
                       {email_confidence} AS email_confidence,
                       {linkedin_verified} AS linkedin_verified,
                       {is_decision_maker} AS is_decision_maker,
                       {contact_priority} AS contact_priority,
                       {enrichment_status} AS enrichment_status
                FROM lead_contacts
                WHERE company_id = %s
                ORDER BY contact_priority DESC, is_decision_maker DESC, created_at ASC
                """.format(**contact_opt),
                (lead_id,),
            )
            contacts_rows = cur.fetchall()

            # Timeline: incorporation, LEI events, officer appointments, RM actions
            cur.execute(
                """
                SELECT event_date, event_type, event_label, event_meta
                FROM (
                    -- Incorporation
                    SELECT
                        incorporation_date      AS event_date,
                        'incorporation'         AS event_type,
                        'Company incorporated'  AS event_label,
                        json_build_object(
                            'jurisdiction', jurisdiction,
                            'entity_type',  entity_type
                        )::text                 AS event_meta
                    FROM companies WHERE id = %(cid)s AND incorporation_date IS NOT NULL

                    UNION ALL

                    -- LEI registration
                    SELECT
                        registered_on           AS event_date,
                        'lei'                   AS event_type,
                        'LEI registered'        AS event_label,
                        json_build_object(
                            'lei_code', lei_code,
                            'status',   registration_status
                        )::text                 AS event_meta
                    FROM lei_records WHERE company_id = %(cid)s AND registered_on IS NOT NULL

                    UNION ALL

                    -- Officer appointments
                    SELECT
                        appointed_on            AS event_date,
                        'officer_appointed'     AS event_type,
                        officer_name || ' appointed as ' || COALESCE(role, 'officer') AS event_label,
                        json_build_object('role', role)::text AS event_meta
                    FROM company_officers
                    WHERE company_id = %(cid)s AND appointed_on IS NOT NULL

                    UNION ALL

                    -- Officer resignations
                    SELECT
                        resigned_on             AS event_date,
                        'officer_resigned'      AS event_type,
                        officer_name || ' resigned'  AS event_label,
                        json_build_object('role', role)::text AS event_meta
                    FROM company_officers
                    WHERE company_id = %(cid)s AND resigned_on IS NOT NULL

                    UNION ALL

                    -- RM status changes from audit log
                    SELECT
                        created_at::date        AS event_date,
                        'rm_action'             AS event_type,
                        action                  AS event_label,
                        COALESCE(new_value::text, '') AS event_meta
                    FROM audit_log
                    WHERE entity_type = 'company' AND entity_id = %(cid)s
                ) t
                ORDER BY event_date DESC NULLS LAST
                LIMIT 30
                """,
                {"cid": lead_id},
            )
            timeline_rows = cur.fetchall()

            route_opt = _route_optional_exprs(cur)
            cur.execute(
                f"""
                SELECT id, contactability_bucket, best_route_type, best_route_value,
                       route_candidate_id, rationale, evidence_summary, missing_data,
                       next_action, confidence, generated_by, generated_at,
                       reviewed_by, reviewed_at, status,
                       NULL::text AS secondary_contact_route,
                       {route_opt["route_source_url"]},
                       {route_opt["route_source_label"]},
                       {route_opt["route_source_type"]},
                       'system_detected'::text AS route_entry_method,
                       {route_opt["route_last_checked_at"]}
                FROM route_recommendations
                WHERE lead_id = %s
                ORDER BY generated_at DESC
                LIMIT 1
                """,
                (lead_id,),
            )
            route_recommendation_row = cur.fetchone()

            cur.execute(
                """
                SELECT id, status, match_type, match_strength, evidence, source
                FROM introducer_matches
                WHERE lead_id = %s AND status IN ('pending', 'accepted')
                ORDER BY
                    CASE match_strength WHEN 'high' THEN 0 WHEN 'medium' THEN 1 ELSE 2 END,
                    created_at DESC
                """,
                (lead_id,),
            )
            introducer_match_rows = cur.fetchall()

    score = {
        "score": row[10] if row[10] is not None else 0,
        "tier": row[11] if row[11] is not None else "LOW",
        "reason_codes": row[12] if row[12] is not None else [],
        "reason_summary": row[13] if row[13] is not None else "No signals matched.",
        "scoring_version": row[14] if row[14] is not None else SCORING_VERSION,
    }
    action = {
        "assigned_to": row[15],
        "status": normalize_status(row[16]) or "new",
        "notes": row[17] or "",
        "contacted_at": row[18],
        "follow_up_at": row[19],
        "next_action": row[20] or "",
        "next_action_due_date": row[21],
        "feedback": row[22] or "",
        "feedback_note": row[23] or "",
    }
    audit_rendered = [
        {
            "entity_type": item[0],
            "entity_id": item[1],
            "action": item[2],
            "actor": item[3],
            "old_value": item[4],
            "new_value": item[5],
            "created_at": item[6],
        }
        for item in audit_rows
    ]
    last_activity = audit_rendered[0]["created_at"] if audit_rendered else None

    lei = None
    if lei_row:
        registered_on = lei_row[3]
        days_ago = (
            (date.today() - registered_on).days
            if registered_on else None
        )
        lei = {
            "lei_code": lei_row[0],
            "entity_status": lei_row[1],
            "registration_status": lei_row[2],
            "registered_on": registered_on,
            "days_ago": days_ago,
            "fresh": days_ago is not None and days_ago <= 90,
            "lou_code": lei_row[4],
            "gleif_url": lei_row[5],
            "last_seen": lei_row[6],
        }

    officers = [
        {
            "name": r[0],
            "role": r[1],
            "appointed_on": r[2],
            "resigned_on": r[3],
            "nationality": r[4],
            "country_of_residence": r[5],
        }
        for r in officers_rows
    ]
    pscs = [
        {
            "name": r[0],
            "kind": r[1],
            "nationality": r[2],
            "country_of_residence": r[3],
            "natures_of_control": r[4] or [],
            "notified_on": r[5],
            "ceased_on": r[6],
        }
        for r in pscs_rows
    ]
    contacts = [
        {
            "id": r[0],
            "name": r[1],
            "role": r[2] or "",
            "email": r[3] or "",
            "phone": r[4] or "",
            "linkedin_url": r[5] or "",
            "source": r[6] or "",
            "notes": r[7] or "",
            "created_at": r[8],
            "email_confidence": r[9] or "",
            "linkedin_verified": r[10] or False,
            "is_decision_maker": r[11] or False,
            "contact_priority": r[12] or 0,
            "enrichment_status": r[13] or "manual",
        }
        for r in contacts_rows
    ]

    _TIMELINE_ICONS = {
        "incorporation":    "🏢",
        "lei":              "🔑",
        "officer_appointed":"👤",
        "officer_resigned": "👋",
        "rm_action":        "📋",
    }
    _TIMELINE_COLORS = {
        "incorporation":    "#1d4ed8",
        "lei":              "#15803d",
        "officer_appointed":"#6b7280",
        "officer_resigned": "#dc2626",
        "rm_action":        "#7c3aed",
    }
    timeline = [
        {
            "date": r[0],
            "type": r[1],
            "label": r[2],
            "icon": _TIMELINE_ICONS.get(r[1], "•"),
            "color": _TIMELINE_COLORS.get(r[1], "#6b7280"),
        }
        for r in timeline_rows
    ]

    route_intelligence = None
    if route_recommendation_row:
        route_intelligence = {
            "id": route_recommendation_row[0],
            "contactability_bucket": route_recommendation_row[1],
            "contactability_label": CONTACTABILITY_LABELS.get(
                route_recommendation_row[1] or "", ""
            ),
            "contactability_status": contactability_status(route_recommendation_row[1]),
            "contactability_status_label": contactability_status_label(
                route_recommendation_row[1]
            ),
            "decision": contactability_decision(route_recommendation_row[1]),
            "suggested_opener": suggested_opener(
                company_name=row[1],
                entity_type=row[3],
                jurisdiction=row[2],
                contactability_bucket=route_recommendation_row[1],
                best_route_value=route_recommendation_row[3],
            ),
            "best_route_type": route_recommendation_row[2],
            "best_route_value": route_recommendation_row[3],
            "route_candidate_id": route_recommendation_row[4],
            "rationale": route_recommendation_row[5],
            "evidence_summary": route_recommendation_row[6] or [],
            "missing_data": route_recommendation_row[7] or [],
            "next_action": route_recommendation_row[8],
            "confidence": route_recommendation_row[9],
            "generated_by": route_recommendation_row[10],
            "generated_at": route_recommendation_row[11],
            "reviewed_by": route_recommendation_row[12],
            "reviewed_at": route_recommendation_row[13],
            "status": route_recommendation_row[14],
            "secondary_contact_route": route_recommendation_row[15],
            "route_source_url": sanitize_external_url(route_recommendation_row[16]),
            "route_source_label": route_recommendation_row[17],
            "route_source_type": route_recommendation_row[18],
            "route_entry_method": route_recommendation_row[19] or "system_detected",
            "route_last_checked_at": route_recommendation_row[20],
        }
        route_intelligence.update(
            lead_readiness(
                recommendation={
                    "contactability_bucket": route_intelligence["contactability_bucket"],
                    "confidence": route_intelligence["confidence"],
                    "evidence_summary": route_intelligence["evidence_summary"],
                    "next_action": route_intelligence["next_action"],
                },
                tier=score["tier"],
            )
        )
        # RM-facing plain-language fields (one consistent status everywhere).
        route_intelligence["rm_status"] = rm_status(
            readiness=route_intelligence["readiness"],
            contactability_bucket=route_intelligence["contactability_bucket"],
        )
        route_intelligence["rm_status_label"] = rm_status_label(
            readiness=route_intelligence["readiness"],
            contactability_bucket=route_intelligence["contactability_bucket"],
        )
        route_intelligence["source_reliability_label"] = source_reliability_label(
            route_intelligence["confidence"]
        )
        route_intelligence["best_route_label"] = best_route_label(
            route_intelligence["best_route_type"], route_intelligence["best_route_value"]
        )

    introducer_matches = [
        {
            "id": r[0],
            "status": r[1],
            "match_type": r[2],
            "match_strength": r[3],
            "evidence": r[4],
            "source": r[5],
        }
        for r in introducer_match_rows
    ]

    # "Why now" — time-sensitive freshness signals for the action panel, derived
    # only from existing score reason codes and LEI recency (no new fetching).
    why_now_bits = []
    if "FRESH_LEI" in score["reason_codes"]:
        why_now_bits.append("Fresh LEI registration")
    if "RECENTLY_INCORPORATED" in score["reason_codes"]:
        why_now_bits.append("Recently incorporated entity")
    if (
        lei
        and lei.get("fresh")
        and lei.get("days_ago") is not None
        and "Fresh LEI registration" not in why_now_bits
    ):
        why_now_bits.append(f"LEI registered {lei['days_ago']} days ago")
    why_now = " · ".join(dict.fromkeys(why_now_bits)) or None

    return templates.TemplateResponse(
        request=request,
        name="lead_detail.html",
        context={
            "lead": {
                "id": row[0],
                "company_name": row[1],
                "jurisdiction": row[2],
                "entity_type": _format_entity_type(row[3]),
                "incorporation_date": row[4],
                "registered_address": row[5],
                "source_system": row[6],
                "source_ref": row[7],
                "verify_url": sanitize_external_url(row[8]),
                "website": sanitize_external_url(row[9]),
            },
            "score": score,
            "action": action,
            "audit_rows": audit_rendered,
            "last_activity": last_activity,
            "lei": lei,
            "officers": officers,
            "pscs": pscs,
            "contacts": contacts,
            "rm_names": RM_NAMES,
            "statuses": _STATUS_OPTIONS,
            "saved": False,
            "actor_names": ACTOR_NAMES,
            "current_actor": (_read_actor(request) or ""),
            "signal_details": SIGNAL_DETAILS,
            "timeline": timeline,
            "route_intelligence": route_intelligence,
            "introducer_matches": introducer_matches,
            "why_now": why_now,
        },
    )


@app.post("/leads/{lead_id}/action", response_class=HTMLResponse)
@write_guard_required
def lead_action(
    request: Request,
    lead_id: UUID,
    assigned_to: str = Form(""),
    status: str = Form("new"),
    notes: str = Form(""),
    contacted_at: str = Form(""),
    follow_up_at: str = Form(""),
):
    status_canonical = _canonical_status_or_422(status)

    # Auto follow-up: if contacted_at is provided and follow_up_at is blank,
    # propose contacted_at + 7 days. Never overwrites a user-entered date.
    effective_follow_up = follow_up_at
    if contacted_at and not follow_up_at:
        try:
            contacted_dt = datetime.strptime(contacted_at, "%Y-%m-%d").date()
            effective_follow_up = (contacted_dt + timedelta(days=7)).isoformat()
        except ValueError:
            pass

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT assigned_to, status, notes, contacted_at, follow_up_at FROM rm_actions WHERE company_id = %s",
                (lead_id,),
            )
            existing = cur.fetchone()
            # If existing follow_up_at already set and user submitted blank, preserve it.
            if existing and existing[4] and not follow_up_at:
                effective_follow_up = existing[4].isoformat()
            old_value = None if existing is None else {
                "assigned_to": existing[0],
                "status": existing[1],
                "notes": existing[2],
                "contacted_at": existing[3].isoformat() if existing[3] else None,
                "follow_up_at": existing[4].isoformat() if existing[4] else None,
            }

            cur.execute(
                """
                INSERT INTO rm_actions (company_id, assigned_to, status, notes, contacted_at, follow_up_at)
                VALUES (%s, %s, %s, %s, NULLIF(%s, '')::date, NULLIF(%s, '')::date)
                ON CONFLICT (company_id) DO UPDATE SET
                    assigned_to = EXCLUDED.assigned_to,
                    status = EXCLUDED.status,
                    notes = EXCLUDED.notes,
                    contacted_at = EXCLUDED.contacted_at,
                    follow_up_at = EXCLUDED.follow_up_at,
                    updated_at = NOW()
                """,
                (
                    lead_id,
                    assigned_to or None,
                    status_canonical,
                    notes or None,
                    contacted_at,
                    effective_follow_up,
                ),
            )

            new_value = {
                "assigned_to": assigned_to or None,
                "status": status_canonical,
                "notes": notes or None,
                "contacted_at": contacted_at or None,
                "follow_up_at": follow_up_at or None,
            }

            cur.execute(
                """
                INSERT INTO audit_log (entity_type, entity_id, action, actor, old_value, new_value, ip_address)
                VALUES ('company', %s, 'rm_action_updated', %s, %s, %s, NULL)
                """,
                (lead_id, (_read_actor(request) or "unknown"), Jsonb(old_value), Jsonb(new_value)),
            )
            conn.commit()

    return _render_action_panel(
        lead_id,
        assigned_to or "",
        status_canonical,
        notes,
        None if not contacted_at else datetime.strptime(contacted_at, "%Y-%m-%d"),
        None if not effective_follow_up else datetime.strptime(effective_follow_up, "%Y-%m-%d"),
        saved=True,
    )


_FEEDBACK_VALUES = {
    "useful",
    "wrong_contact",
    "not_relevant",
    "duplicate",
    "contacted",
    "meeting_booked",
    "won",
    "lost",
}


@app.post("/leads/{lead_id}/feedback", response_class=HTMLResponse)
@write_guard_required
def lead_feedback(
    request: Request,
    lead_id: UUID,
    feedback: str = Form(""),
    feedback_note: str = Form(""),
    next_action: str = Form(""),
    next_action_due_date: str = Form(""),
):
    """Record RM feedback and next-action on the lead's rm_actions row.

    Additive to the existing RM action flow: it touches only the feedback /
    next-action columns and writes an audit_log entry. Status, notes, and
    follow-up handled by /action are left untouched.
    """
    feedback_value = (feedback or "").strip()
    if feedback_value and feedback_value not in _FEEDBACK_VALUES:
        raise HTTPException(status_code=422, detail="Unknown feedback value")
    actor = _read_actor(request) or "unknown"

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT feedback, feedback_note, next_action, next_action_due_date "
                "FROM rm_actions WHERE company_id = %s",
                (lead_id,),
            )
            existing = cur.fetchone()
            old_value = None if existing is None else {
                "feedback": existing[0],
                "feedback_note": existing[1],
                "next_action": existing[2],
                "next_action_due_date": (
                    existing[3].isoformat() if existing[3] else None
                ),
            }

            cur.execute(
                """
                INSERT INTO rm_actions
                    (company_id, feedback, feedback_note, next_action, next_action_due_date)
                VALUES (
                    %s, NULLIF(%s, ''), NULLIF(%s, ''), NULLIF(%s, ''),
                    NULLIF(%s, '')::date
                )
                ON CONFLICT (company_id) DO UPDATE SET
                    feedback = EXCLUDED.feedback,
                    feedback_note = EXCLUDED.feedback_note,
                    next_action = EXCLUDED.next_action,
                    next_action_due_date = EXCLUDED.next_action_due_date,
                    updated_at = NOW()
                """,
                (
                    lead_id,
                    feedback_value,
                    feedback_note,
                    next_action,
                    next_action_due_date,
                ),
            )

            new_value = {
                "feedback": feedback_value or None,
                "feedback_note": feedback_note or None,
                "next_action": next_action or None,
                "next_action_due_date": next_action_due_date or None,
            }
            cur.execute(
                """
                INSERT INTO audit_log
                    (entity_type, entity_id, action, actor, old_value, new_value, ip_address)
                VALUES ('company', %s, 'rm_feedback_updated', %s, %s, %s, NULL)
                """,
                (lead_id, actor, Jsonb(old_value), Jsonb(new_value)),
            )
            conn.commit()

    return RedirectResponse(url=f"/leads/{lead_id}", status_code=303)


@app.post("/leads/{lead_id}/contacts", response_class=HTMLResponse)
@write_guard_required
def add_lead_contact(
    request: Request,
    lead_id: UUID,
    name: str = Form(...),
    role: str = Form(""),
    email: str = Form(""),
    phone: str = Form(""),
    linkedin_url: str = Form(""),
    source: str = Form("manual"),
    notes: str = Form(""),
):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO lead_contacts
                    (company_id, name, role, email, phone, linkedin_url, source, notes)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (str(lead_id), name.strip(), role.strip(), email.strip(),
                 phone.strip(), linkedin_url.strip(), source, notes.strip()),
            )
        conn.commit()
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url=f"/leads/{lead_id}", status_code=303)


@app.post("/leads/{lead_id}/contacts/{contact_id}/delete", response_class=HTMLResponse)
@write_guard_required
def delete_lead_contact(request: Request, lead_id: UUID, contact_id: UUID):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "DELETE FROM lead_contacts WHERE id = %s AND company_id = %s",
                (str(contact_id), str(lead_id)),
            )
        conn.commit()
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url=f"/leads/{lead_id}", status_code=303)


# --- Inline assign/status endpoint for queue HTMX ---
@app.post("/leads/{lead_id}/assign", response_class=HTMLResponse)
@write_guard_required
def lead_assign(
        request: Request,
        lead_id: UUID,
        assigned_to: str = Form(""),
        status: str = Form("new"),
):
        status_canonical = _canonical_status_or_422(status)
        with get_conn() as conn:
                with conn.cursor() as cur:
                        cur.execute(
                                """
                                INSERT INTO rm_actions (company_id, assigned_to, status)
                                VALUES (%s, %s, %s)
                                ON CONFLICT (company_id) DO UPDATE SET
                                        assigned_to = EXCLUDED.assigned_to,
                                        status = EXCLUDED.status,
                                        updated_at = NOW()
                                """,
                                (lead_id, assigned_to or None, status_canonical),
                        )
                        cur.execute(
                                """
                                INSERT INTO audit_log
                                    (entity_type, entity_id, action, actor, new_value)
                                VALUES ('company', %s, 'quick_assign', %s, %s)
                                """,
                                (
                                    lead_id,
                                    (_read_actor(request) or "unknown"),
                                    Jsonb({"assigned_to": assigned_to or None, "status": status_canonical}),
                                ),
                        )
                        conn.commit()

                with conn.cursor() as cur:
                        cur.execute(
                                """
                                SELECT c.id, c.company_name, c.jurisdiction, c.entity_type,
                                             c.incorporation_date, c.verify_url,
                                             qs.priority_score, qs.tier, qs.reason_summary,
                                             ra.assigned_to, ra.status, qs.refreshed_at
                                FROM queue_snapshot qs
                                JOIN companies c ON c.id = qs.canonical_company_id
                                LEFT JOIN rm_actions ra ON ra.company_id = c.id
                                WHERE c.id = %s
                                """,
                                (lead_id,),
                        )
                        row = cur.fetchone()

        if row is None:
                return HTMLResponse("<tr><td colspan='8'>Not found</td></tr>", status_code=404)

        r = {
                "id": row[0], "company_name": row[1], "jurisdiction": row[2],
                "entity_type": _format_entity_type(row[3]),
                "incorporation_date": row[4], "verify_url": sanitize_external_url(row[5]),
                "priority_score": row[6], "tier": row[7], "reason_summary": row[8],
                "assigned_to": row[9], "status": normalize_status(row[10]) or "new",
        }

        score_pct = r["priority_score"]
        score_color = "#059669" if score_pct >= 70 else "#d97706" if score_pct >= 40 else "#d1d5db"
        assigned_opts = "".join(
                f'<option value="{nm}" {"selected" if nm == (r["assigned_to"] or "") else ""}>{nm}</option>'
                for nm in RM_NAMES
        )
        status_opts = "".join(
                f'<option value="{s["value"]}" {"selected" if s["value"] == (r["status"] or "new") else ""}>{s["label"]}</option>'
                for s in _STATUS_OPTIONS
        )
        verify = (
            f'<a href="{escape(r["verify_url"], quote=True)}" target="_blank" style="font-size:11px">↗</a>'
            if r["verify_url"]
            else ""
        )

        return HTMLResponse(f"""
        <tr>
            <td><a href="/leads/{r['id']}">{escape(r['company_name'])}</a></td>
            <td>{escape(r['jurisdiction'])}</td>
            <td>{escape(r['entity_type'])}</td>
            <td>{r['incorporation_date'] or '—'}</td>
            <td>
                <div style=\"min-width:44px\">
                    <div style=\"height:3px;border-radius:2px;margin-bottom:3px;width:{score_pct}%;background:{score_color}\"></div>
                    <strong>{score_pct}</strong>
                </div>
            </td>
            <td><span class=\"tier tier-{r['tier']}\">{r['tier']}</span></td>
            <td>
                <form hx-post=\"/leads/{r['id']}/assign\" hx-target=\"closest tr\" hx-swap=\"outerHTML\" style=\"margin:0\">
                    <input type=\"hidden\" name=\"status\" value=\"{escape(r['status'] or 'new')}\">
                    <select name=\"assigned_to\" onchange=\"this.form.requestSubmit()\"
                        style=\"font-size:11px;padding:2px 4px;border:1px solid #ddd;border-radius:4px;max-width:90px\">
                        <option value=\"\">—</option>
                        {assigned_opts}
                    </select>
                </form>
            </td>
            <td>
                <form hx-post=\"/leads/{r['id']}/assign\" hx-target=\"closest tr\" hx-swap=\"outerHTML\" style=\"margin:0\">
                    <input type=\"hidden\" name=\"assigned_to\" value=\"{escape(r['assigned_to'] or '')}\">
                    <select name=\"status\" onchange=\"this.form.requestSubmit()\"
                        style=\"font-size:11px;padding:2px 4px;border:1px solid #ddd;border-radius:4px;max-width:110px\">
                        {status_opts}
                    </select>
                </form>
            </td>
            <td>{verify}</td>
        </tr>""")


@app.post("/leads/{lead_id}/publish", response_class=HTMLResponse)
@write_guard_required
def publish_lead_to_team(request: Request, lead_id: UUID) -> RedirectResponse:
    actor = _read_actor(request)
    if not _is_admin_actor(actor):
        raise HTTPException(status_code=403, detail="Admin access required")

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT rr.id
                FROM route_recommendations rr
                WHERE rr.lead_id = %s
                  AND rr.status <> 'superseded'
                  AND rr.contactability_bucket = ANY(%s)
                  AND NULLIF(TRIM(rr.best_route_value), '') IS NOT NULL
                  AND NULLIF(TRIM(rr.next_action), '') IS NOT NULL
                  AND jsonb_array_length(COALESCE(rr.evidence_summary, '[]'::jsonb)) > 0
                ORDER BY rr.generated_at DESC NULLS LAST
                LIMIT 1
                """,
                (lead_id, list(_ELIGIBLE_ROUTE_BUCKETS)),
            )
            if cur.fetchone() is None:
                raise HTTPException(
                    status_code=409,
                    detail="Lead needs a compliant contact route before publishing",
                )

            cur.execute(
                """
                INSERT INTO rm_actions (company_id, status)
                VALUES (%s, 'sent_to_team')
                ON CONFLICT (company_id) DO UPDATE SET
                    status = 'sent_to_team',
                    updated_at = NOW()
                """,
                (lead_id,),
            )
            cur.execute(
                """
                INSERT INTO audit_log
                    (entity_type, entity_id, action, actor, new_value)
                VALUES ('company', %s, 'published_to_team', %s, %s)
                """,
                (lead_id, actor, Jsonb({"status": "sent_to_team"})),
            )
        conn.commit()

    return RedirectResponse(url="/?view=admin&tab=enriched", status_code=303)


def _review_route_recommendation(
    request: Request, lead_id: UUID, recommendation_id: UUID, new_status: str
) -> RedirectResponse:
    """Accept/reject a route recommendation and record an auditable review.

    Updates only the route-intelligence review surface — never scoring, RM
    actions, or outreach. Mirrors the recommendation decision onto any linked
    introducer match and writes a single audit_log entry. The actor check runs
    before any DB access so unauthenticated calls never touch the database.
    """
    actor = _read_actor(request)
    if not actor:
        raise HTTPException(
            status_code=400, detail="Actor required to review a route recommendation"
        )
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT status, route_candidate_id, contactability_bucket,
                       best_route_type, best_route_value, evidence_summary
                FROM route_recommendations
                WHERE id = %s AND lead_id = %s
                """,
                (recommendation_id, lead_id),
            )
            row = cur.fetchone()
            if row is None:
                raise HTTPException(
                    status_code=404, detail="Route recommendation not found"
                )
            current_status, introducer_id = row[0], row[1]

            cur.execute(
                """
                UPDATE route_recommendations
                SET status = %s, reviewed_by = %s, reviewed_at = NOW()
                WHERE id = %s AND lead_id = %s
                """,
                (new_status, actor, recommendation_id, lead_id),
            )

            if introducer_id is not None:
                cur.execute(
                    """
                    UPDATE introducer_matches
                    SET status = %s, reviewed_by = %s, reviewed_at = NOW()
                    WHERE lead_id = %s AND introducer_id = %s
                    """,
                    (new_status, actor, lead_id, introducer_id),
                )

            cur.execute(
                """
                INSERT INTO audit_log
                    (entity_type, entity_id, action, actor, old_value, new_value)
                VALUES ('company', %s, %s, %s, %s, %s)
                """,
                (
                    lead_id,
                    f"route_recommendation_{new_status}",
                    actor,
                    Jsonb({"recommendation_id": str(recommendation_id), "status": current_status}),
                    Jsonb({"recommendation_id": str(recommendation_id), "status": new_status}),
                ),
            )
        conn.commit()
    return RedirectResponse(url=f"/leads/{lead_id}", status_code=303)


@app.post(
    "/leads/{lead_id}/route-recommendations/{recommendation_id}/accept",
    response_class=HTMLResponse,
)
def accept_route_recommendation(
    request: Request, lead_id: UUID, recommendation_id: UUID
):
    return _review_route_recommendation(request, lead_id, recommendation_id, "accepted")


@app.post(
    "/leads/{lead_id}/route-recommendations/{recommendation_id}/reject",
    response_class=HTMLResponse,
)
def reject_route_recommendation(
    request: Request, lead_id: UUID, recommendation_id: UUID
):
    return _review_route_recommendation(request, lead_id, recommendation_id, "rejected")


@app.get("/upload", response_class=HTMLResponse)
def upload_form(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="upload.html",
        context={"preview": None, "actor_names": ACTOR_NAMES, "current_actor": (_read_actor(request) or "")},
    )


@app.post("/upload", response_class=HTMLResponse)
@write_guard_required
def upload_preview(request: Request, file: UploadFile = File(...)):
    file_bytes = file.file.read()
    columns, rows, validation_errors = _parse_upload_csv(file_bytes)

    preview = {
        "id": None,
        "row_count": len(rows),
        "columns": columns,
        "rows": rows,
        "validation_errors": validation_errors,
    }

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO pending_uploads (filename, uploaded_by, row_count, parsed_rows, validation_errors, status)
                VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING id
                """,
                (
                    file.filename or "upload.csv",
                    (_read_actor(request) or "unknown"),
                    len(rows),
                    Jsonb(rows),
                    Jsonb(validation_errors),
                    "pending" if not validation_errors else "rejected",
                ),
            )
            preview["id"] = cur.fetchone()[0]
            conn.commit()

    return templates.TemplateResponse(
        request=request,
        name="upload.html",
        context={"preview": preview, "actor_names": ACTOR_NAMES, "current_actor": (_read_actor(request) or "")},
    )


@app.post("/upload/{upload_id}/confirm")
@write_guard_required
def upload_confirm(request: Request, upload_id: UUID):
    with get_conn() as conn:
        with conn.cursor() as cur:
            # Atomically claim the upload: only succeeds if status is currently 'pending'.
            # Concurrent requests will find no row to update and receive a 409.
            cur.execute(
                """
                UPDATE pending_uploads
                SET status = 'confirmed'
                WHERE id = %s AND status = 'pending'
                RETURNING parsed_rows, validation_errors
                """,
                (upload_id,),
            )
            claimed = cur.fetchone()
            if claimed is None:
                # Either not found or already confirmed/rejected
                cur.execute(
                    "SELECT 1 FROM pending_uploads WHERE id = %s",
                    (upload_id,),
                )
                exists = cur.fetchone()
                if exists is None:
                    raise HTTPException(status_code=404, detail="Upload not found")
                raise HTTPException(status_code=409, detail="Upload already confirmed or rejected")

            parsed_rows, validation_errors = claimed
            if validation_errors:
                # Roll back the status change — this upload has errors and cannot be confirmed
                conn.rollback()
                raise HTTPException(status_code=400, detail="Upload has validation errors and cannot be confirmed")

            parsed_rows = parsed_rows or []
            inserted = 0
            skipped_duplicates = 0
            for index, row in enumerate(parsed_rows, start=1):
                company_name = (row.get("company_name") or "").strip()
                jurisdiction = (row.get("jurisdiction") or "").strip()
                entity_type = (row.get("entity_type") or "").strip() or None
                website = (row.get("website") or "").strip() or None
                normalised_name = re.sub(r"\s+", " ", re.sub(r"[^a-z0-9 ]", "", company_name.lower())).strip()
                source_ref = f"upload:{upload_id}:{index}"

                cur.execute(
                    """
                    SELECT id, source_system FROM companies
                    WHERE normalised_name = %s AND jurisdiction = %s
                    LIMIT 1
                    """,
                    (normalised_name, jurisdiction),
                )
                existing = cur.fetchone()
                if existing is not None:
                    skipped_duplicates += 1
                    logger.warning(
                        "upload_dedup_skipped",
                        extra={
                            "upload_id": str(upload_id),
                            "row_index": index,
                            "company_name": company_name,
                            "jurisdiction": jurisdiction,
                            "existing_company_id": str(existing[0]),
                            "existing_source_system": existing[1],
                        },
                    )
                    continue

                inserted += 1
                cur.execute(
                    """
                    INSERT INTO companies (
                        source_system, source_ref, company_name, normalised_name,
                        jurisdiction, entity_type, incorporation_date,
                        registered_address, sic_codes, website, verify_url, raw_data
                    ) VALUES (
                        %s, %s, %s, %s,
                        %s, %s, %s,
                        %s, %s, %s, %s, %s
                    )
                    ON CONFLICT (source_system, source_ref) DO UPDATE SET
                        company_name = EXCLUDED.company_name,
                        normalised_name = EXCLUDED.normalised_name,
                        jurisdiction = EXCLUDED.jurisdiction,
                        entity_type = EXCLUDED.entity_type,
                        incorporation_date = EXCLUDED.incorporation_date,
                        registered_address = EXCLUDED.registered_address,
                        sic_codes = EXCLUDED.sic_codes,
                        website = EXCLUDED.website,
                        verify_url = EXCLUDED.verify_url,
                        raw_data = EXCLUDED.raw_data,
                        updated_at = NOW()
                    """,
                    (
                        "manual_upload",
                        source_ref,
                        company_name,
                        normalised_name,
                        jurisdiction,
                        entity_type,
                        None,
                        None,
                        [],
                        website,
                        None,
                        Jsonb(row),
                    ),
                )

            logger.info(
                "upload_confirmed",
                extra={
                    "upload_id": str(upload_id),
                    "rows_total": len(parsed_rows),
                    "rows_inserted": inserted,
                    "rows_skipped_duplicate": skipped_duplicates,
                },
            )
            conn.commit()

    return RedirectResponse(url="/", status_code=303)


_AUDIT_ACTION_LABELS = {
    "rm_action_updated": "Lead updated",
    "rm_feedback_updated": "RM feedback recorded",
    "route_recommendation_accepted": "Route accepted",
    "route_recommendation_rejected": "Route rejected",
    "quick_assign": "Quick assign",
    "lead_status_changed": "Status changed",
    "lead_assigned": "Reassigned",
    "score_recalculated": "Score recalculated",
    "introducer_updated": "Introducer updated",
}

_AUDIT_FIELD_LABELS = {
    "assigned_to": "Assigned to",
    "status": "Status",
    "notes": "Notes",
    "contacted_at": "Contacted",
    "follow_up_at": "Follow-up",
    "tier": "Tier",
    "priority_score": "Score",
}


def _audit_action_label(action: str | None) -> str:
    if not action:
        return "—"
    return _AUDIT_ACTION_LABELS.get(action, action.replace("_", " ").capitalize())


def _audit_field_label(field: str) -> str:
    return _AUDIT_FIELD_LABELS.get(field, field.replace("_", " ").capitalize())


def _format_audit_value(value) -> str:
    if value is None or value == "":
        return "—"
    if isinstance(value, str):
        return value
    return str(value)


def _audit_changes(old, new) -> list[dict]:
    """Return list of changed fields {label, old, new}. Falls back to a single
    raw 'new' row when payloads aren't dict-compatible."""
    if not isinstance(new, dict):
        if new is None and old is None:
            return []
        return [{"label": "Value", "old": _format_audit_value(old), "new": _format_audit_value(new)}]
    old_dict = old if isinstance(old, dict) else {}
    changes = []
    # union of keys preserves all transitions, including new keys
    for key in new.keys():
        old_v = old_dict.get(key)
        new_v = new.get(key)
        if old_v == new_v:
            continue
        changes.append({
            "label": _audit_field_label(key),
            "old": _format_audit_value(old_v),
            "new": _format_audit_value(new_v),
        })
    # if nothing changed but it's still a meaningful event, show a compact summary
    if not changes:
        for key, val in new.items():
            if val in (None, ""):
                continue
            changes.append({
                "label": _audit_field_label(key),
                "old": None,
                "new": _format_audit_value(val),
            })
    return changes


def _humanize_actor(actor: str | None) -> str:
    if not actor or actor.lower() in ("unknown", ""):
        return "Unknown"
    if actor.lower() == "system":
        return "System"
    return actor


def _relative_time(ts: datetime | None) -> str:
    if not ts:
        return ""
    now = datetime.now(timezone.utc)
    if ts.tzinfo is None:
        ts = ts.replace(tzinfo=timezone.utc)
    delta = now - ts
    secs = int(delta.total_seconds())
    if secs < 60:
        return "just now"
    if secs < 3600:
        return f"{secs // 60}m ago"
    if secs < 86400:
        return f"{secs // 3600}h ago"
    if secs < 7 * 86400:
        return f"{secs // 86400}d ago"
    return ts.strftime("%d %b %Y")


@app.get("/audit", response_class=HTMLResponse)
def audit_log(request: Request):
    filters = {
        "actor": request.query_params.get("actor", ""),
        "action": request.query_params.get("action", ""),
        "date_from": request.query_params.get("date_from", ""),
        "date_to": request.query_params.get("date_to", ""),
    }
    try:
        page = max(int(request.query_params.get("page", 1)), 1)
    except ValueError:
        page = 1
    page_size = 50

    where = ["1=1"]
    params: list[object] = []
    if filters["actor"]:
        where.append("a.actor = %s")
        params.append(filters["actor"])
    if filters["action"]:
        where.append("a.action = %s")
        params.append(filters["action"])
    if filters["date_from"]:
        where.append("a.created_at >= %s")
        params.append(filters["date_from"])
    if filters["date_to"]:
        where.append("a.created_at < (%s::date + INTERVAL '1 day')")
        params.append(filters["date_to"])

    where_sql = " AND ".join(where)

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(f"SELECT COUNT(*) FROM audit_log a WHERE {where_sql}", params)
            total = cur.fetchone()[0]
            cur.execute(
                f"""
                SELECT a.entity_type, a.entity_id, a.action, a.actor,
                       a.old_value, a.new_value, a.created_at,
                       c.company_name
                FROM audit_log a
                LEFT JOIN companies c
                  ON a.entity_type = 'company' AND c.id = a.entity_id
                WHERE {where_sql}
                ORDER BY a.created_at DESC
                LIMIT %s OFFSET %s
                """,
                params + [page_size, (page - 1) * page_size],
            )
            rows = cur.fetchall()

            cur.execute("SELECT DISTINCT actor FROM audit_log WHERE actor IS NOT NULL ORDER BY 1")
            actor_options = [r[0] for r in cur.fetchall()]
            cur.execute("SELECT DISTINCT action FROM audit_log WHERE action IS NOT NULL ORDER BY 1")
            action_options = [r[0] for r in cur.fetchall()]

    total_pages = max((total + page_size - 1) // page_size, 1)
    rendered_rows = []
    for row in rows:
        entity_type, entity_id, action, actor, old_v, new_v, created_at, company_name = row
        rendered_rows.append({
            "entity_type": entity_type,
            "entity_id": entity_id,
            "entity_url": f"/leads/{entity_id}" if entity_type == "company" else None,
            "entity_label": company_name or (str(entity_id)[:8] if entity_id else "—"),
            "action": action,
            "action_label": _audit_action_label(action),
            "actor": _humanize_actor(actor),
            "actor_raw": actor or "",
            "changes": _audit_changes(old_v, new_v),
            "created_at": created_at,
            "created_at_iso": created_at.isoformat() if created_at else "",
            "created_at_display": created_at.strftime("%d %b %Y %H:%M") if created_at else "",
            "created_at_relative": _relative_time(created_at),
        })

    action_choices = [(a, _audit_action_label(a)) for a in action_options]

    return templates.TemplateResponse(
        request=request,
        name="audit.html",
        context={
            "rows": rendered_rows,
            "total": total,
            "page": page,
            "total_pages": total_pages,
            "actor_names": ACTOR_NAMES,
            "current_actor": (_read_actor(request) or ""),
            "filters": filters,
            "actor_options": actor_options,
            "action_choices": action_choices,
        },
    )
